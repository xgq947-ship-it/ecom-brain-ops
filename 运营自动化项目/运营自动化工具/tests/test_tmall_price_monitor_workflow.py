from __future__ import annotations

import json
from pathlib import Path

import pytest

from core.runtime import WorkflowRunner
from core.runtime.registry import discover_workflow
from core.task_registry import resolve_task

from workflows.tmall_price_monitor import control_price_mapper as mapper
from workflows.tmall_price_monitor import price_compare, steps
from workflows.tmall_price_monitor.report_writer import build_json_payload, write_excel
from workflows.tmall_price_monitor.workflow import build_workflow


# ---------- 测试夹具：合成的猫超 / 聚水潭 CSV ----------

def _write_csv(path: Path, rows: list[list[str]]) -> Path:
    import csv

    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle).writerows(rows)
    return path


@pytest.fixture
def maochao_csv(tmp_path: Path) -> Path:
    return _write_csv(
        tmp_path / "猫超商品列表导出.csv",
        [
            ["商品编码", "商品名称", "条码"],
            ["111", "商品甲", "BC1"],
            ["222", "商品乙", "BC2"],   # 多条码，控价 600
            ["222", "商品乙", "BC3"],   # 多条码，控价 400 → 取 min
            ["333", "商品丙", "BCX"],   # 条码不在聚水潭
            ["444", "商品丁", "BC4"],   # 聚水潭控价为空
        ],
    )


@pytest.fixture
def jst_csv(tmp_path: Path) -> Path:
    return _write_csv(
        tmp_path / "聚水潭商品资料.csv",
        [
            ["商品编码", "商品名称", "淘系控价"],
            ["BC1", "甲SKU", "500"],
            ["BC2", "乙SKU大", "600"],
            ["BC3", "乙SKU小", "400"],
            ["BC4", "丁SKU", ""],
        ],
    )


# ---------- 抓价行构造 ----------

def _scrape(item_id: str, *, price=None, status="ok", title="天猫商品") -> dict:
    return {
        "item_id": item_id,
        "title": title,
        "realtime_price": price,
        "raw_price_text": "¥0",
        "screenshot_path": f"/tmp/{item_id}.png",
        "captured_at": "2026-06-21T00:00:00",
        "capture_status": status,
        "error": None,
    }


def _mapping(item_id: str, *, control=None, status=mapper.STATUS_MATCHED, barcode="BC", name="JST名") -> dict:
    return {
        "item_id": item_id,
        "barcode": barcode,
        "jst_goods_code": barcode,
        "jst_goods_name": name,
        "maochao_name": "猫超名",
        "taoxi_control_price": control,
        "mapping_status": status,
        "matched_barcode_count": 1,
        "all_control_prices": [control] if control is not None else [],
    }


# ---------- 1. 注册 / 别名 ----------

def test_workflow_registers() -> None:
    wf = discover_workflow("tmall_price_monitor")
    assert wf.id == "tmall_price_monitor"
    assert [s.id for s in wf.steps] == [
        "check_inputs",
        "resolve_control_prices",
        "fetch_realtime_prices",
        "compare_prices",
        "notify_login_if_needed",
        "write_outputs",
        "collect_outputs",
    ]


def test_chinese_alias_resolves() -> None:
    assert resolve_task("天猫商品价格监控") == "tmall_price_monitor"
    assert resolve_task("天猫控价监控") == "tmall_price_monitor"


# ---------- 2. 清洗函数 ----------

def test_clean_helpers() -> None:
    assert mapper.clean_item_id(762065566026.0) == "762065566026"
    assert mapper.clean_item_id(" 123.0 ") == "123"
    assert mapper.clean_barcode("AUAMKDK1606") == "AUAMKDK1606"
    assert mapper.clean_price("¥1,299.00 元") == 1299.0
    assert mapper.clean_price("") is None
    assert mapper.clean_price(None) is None
    assert mapper.clean_price(598) == 598.0


# ---------- 3. mapper：四种匹配状态（真实读 CSV） ----------

def test_mapper_status_matrix(maochao_csv: Path, jst_csv: Path) -> None:
    r = mapper.ControlPriceResolver(maochao_path=maochao_csv, jst_path=jst_csv, debug=False)

    m = r.resolve("111")
    assert m["mapping_status"] == mapper.STATUS_MATCHED
    assert m["taoxi_control_price"] == 500.0
    assert m["barcode"] == "BC1"

    # 多条码取最小控价
    m = r.resolve("222")
    assert m["mapping_status"] == mapper.STATUS_MATCHED
    assert m["taoxi_control_price"] == 400.0
    assert m["matched_barcode_count"] == 2
    assert sorted(m["all_control_prices"]) == [400.0, 600.0]

    # 未找到猫超条码
    assert r.resolve("999")["mapping_status"] == mapper.STATUS_NO_BARCODE
    # 未找到聚水潭商品
    assert r.resolve("333")["mapping_status"] == mapper.STATUS_NO_JST
    # 控价为空
    assert r.resolve("444")["mapping_status"] == mapper.STATUS_NO_CONTROL


def test_mapper_missing_file_raises_readable() -> None:
    with pytest.raises(mapper.MappingError):
        mapper.ControlPriceResolver(maochao_path="/no/such/猫超.csv", jst_path="/no/such/jst.csv", debug=False)


# ---------- 4. 输入解析 ----------

def test_resolve_item_ids_and_csv(tmp_path: Path) -> None:
    csv = _write_csv(tmp_path / "items.csv", [["item_id"], ["111"], ["222"], [" 333 "]])
    assert price_compare.load_item_ids_from_csv(csv) == ["111", "222", "333"]
    ids = price_compare.resolve_item_ids(item_id="999", item_ids="111,222", file=csv)
    assert ids == ["111", "222", "999", "333"]  # 去重保序
    with pytest.raises(price_compare.InputError):
        price_compare.resolve_item_ids(item_id=None, item_ids=None, file=None)


# ---------- 5. build_record 状态矩阵 ----------

def test_build_record_matrix() -> None:
    # 低于控价
    r = price_compare.build_record(_scrape("1", price=380.0), _mapping("1", control=409.0))
    assert r["status"] == "低于控价" and r["diff_price"] == -29.0
    # 正常
    r = price_compare.build_record(_scrape("1", price=500.0), _mapping("1", control=409.0))
    assert r["status"] == "正常" and r["diff_price"] == 91.0
    # 未找到猫超条码（未抓价）
    r = price_compare.build_record(None, _mapping("1", status=mapper.STATUS_NO_BARCODE))
    assert r["status"] == "未找到猫超条码" and r["diff_price"] is None
    # 未找到聚水潭商品
    r = price_compare.build_record(None, _mapping("1", status=mapper.STATUS_NO_JST))
    assert r["status"] == "未找到聚水潭商品"
    # 控价为空
    r = price_compare.build_record(None, _mapping("1", status=mapper.STATUS_NO_CONTROL))
    assert r["status"] == "控价为空"
    # 控价匹配成功但抓价失败 → 抓取失败
    r = price_compare.build_record(_scrape("1", status="failed"), _mapping("1", control=409.0))
    assert r["status"] == "抓取失败"
    # 登录/验证码异常
    r = price_compare.build_record(_scrape("1", status="login_required"), _mapping("1", control=409.0))
    assert r["status"] == "登录/验证码异常"
    # 只给商品ID导致活动上下文缺失
    r = price_compare.build_record(_scrape("1", status="price_context_missing"), _mapping("1", control=409.0))
    assert r["status"] == "价格上下文缺失" and r["diff_price"] is None


# ---------- 6. 端到端（真实 mapper + mock 抓价） ----------

def _run(monkeypatch, tmp_path: Path, args, *, dry_run, rows, output_dir):
    seen = []
    sent = []

    def fake_run_ops_json(command, interactive_recovery=None):
        seen.append((list(command), interactive_recovery))
        return {
            "success": True, "platform": "tmall", "command": "price get",
            "data": {"rows": rows, "count": len(rows), "source": "simulated" if dry_run else "page",
                     "simulated": dry_run, "dry_run": dry_run, "artifacts": [], "context_path": "/tmp/c.json"},
        }

    monkeypatch.setattr(steps, "run_ops_json", fake_run_ops_json)

    def forbidden_send_notification(*args, **kwargs):
        sent.append((args, kwargs))
        raise AssertionError("tmall_price_monitor 登录失效不应发送任何通知")

    monkeypatch.setattr(steps, "send_notification", forbidden_send_notification, raising=False)
    monkeypatch.setattr(steps, "DEFAULT_OUTPUT_DIR", output_dir)
    runner = WorkflowRunner(tmp_path)
    run = runner.run(build_workflow(), inputs={"dry_run": dry_run, "args": args}, dry_run=dry_run)
    return run, seen, sent, runner


def _step_outputs(runner, step_id):
    path = runner.last_run_dir / "steps" / f"{step_id}.json"
    return json.loads(path.read_text(encoding="utf-8"))["outputs"]


def test_end_to_end_below_and_unmatched(monkeypatch, tmp_path, maochao_csv, jst_csv) -> None:
    out = tmp_path / "out"
    # 111 控价 500，实时 480 → 低于控价；999 未找到猫超条码（不会被抓价）
    run, seen, _, runner = _run(
        monkeypatch, tmp_path,
        args=["--item-ids", "111,999", "--maochao-file", str(maochao_csv), "--jst-file", str(jst_csv)],
        dry_run=False,
        rows=[_scrape("111", price=480.0)],
        output_dir=out,
    )
    assert run.status == "success"
    # 只对匹配到控价的商品抓价：命令里只含 111
    command, _ = seen[0]
    assert "111" in command and "999" not in ",".join(command)

    cmp_out = _step_outputs(runner, "compare_prices")
    assert cmp_out["summary"].get("低于控价") == 1
    assert cmp_out["summary"].get("未找到猫超条码") == 1
    assert cmp_out["below_control"][0]["item_id"] == "111"
    assert cmp_out["below_control"][0]["taoxi_control_price"] == 500.0

    arts = {a.type for a in run.artifacts}
    assert {"xlsx", "json"} <= arts
    for art in run.artifacts:
        assert Path(art.path).exists()


def test_full_item_url_keeps_marketing_params_for_fetch(monkeypatch, tmp_path, maochao_csv, jst_csv) -> None:
    out = tmp_path / "out"
    item_url = "https://detail.tmall.com/item.htm?id=111&mi_id=abc"
    run, seen, _, runner = _run(
        monkeypatch, tmp_path,
        args=["--item-id", item_url, "--maochao-file", str(maochao_csv), "--jst-file", str(jst_csv)],
        dry_run=False,
        rows=[_scrape("111", price=313.87)],
        output_dir=out,
    )

    assert run.status == "success"
    command, _ = seen[0]
    item_arg = command[command.index("--item-ids") + 1]
    assert item_arg == item_url
    cmp_out = _step_outputs(runner, "compare_prices")
    assert cmp_out["summary"].get("低于控价") == 1


def test_no_matched_skips_fetch(monkeypatch, tmp_path, maochao_csv, jst_csv) -> None:
    out = tmp_path / "out"
    run, seen, _, runner = _run(
        monkeypatch, tmp_path,
        args=["--item-ids", "999", "--maochao-file", str(maochao_csv), "--jst-file", str(jst_csv)],
        dry_run=False, rows=[], output_dir=out,
    )
    assert run.status == "success"
    assert seen == []  # 无匹配项，根本不调用抓价
    cmp_out = _step_outputs(runner, "compare_prices")
    assert cmp_out["summary"].get("未找到猫超条码") == 1


def test_login_failure_does_not_send_notification(monkeypatch, tmp_path, maochao_csv, jst_csv) -> None:
    out = tmp_path / "out"
    run, _, sent, runner = _run(
        monkeypatch, tmp_path,
        args=["--item-ids", "111,222", "--maochao-file", str(maochao_csv), "--jst-file", str(jst_csv)],
        dry_run=False,
        rows=[
            _scrape("111", status="login_required"),
            _scrape("222", status="captcha"),
        ],
        output_dir=out,
    )

    assert run.status == "success"
    assert sent == []
    out = _step_outputs(runner, "notify_login_if_needed")
    assert out["notification"]["sent"] is False
    assert "关闭" in out["notification"]["reason"]
    assert out["affected_count"] == 2
    assert out["affected_item_ids"] == ["111", "222"]


def test_login_failure_dry_run_does_not_send(monkeypatch, tmp_path, maochao_csv, jst_csv) -> None:
    out = tmp_path / "out"
    run, _, sent, runner = _run(
        monkeypatch, tmp_path,
        args=["--item-ids", "111", "--maochao-file", str(maochao_csv), "--jst-file", str(jst_csv), "--dry-run"],
        dry_run=True,
        rows=[_scrape("111", status="login_required")],
        output_dir=out,
    )

    assert run.status == "dry_run_success"
    assert sent == []
    out = _step_outputs(runner, "notify_login_if_needed")
    assert out["notification"]["sent"] is False
    assert "关闭" in out["notification"]["reason"]
    assert out["affected_item_ids"] == ["111"]


# ---------- 7. 产出格式 ----------

def test_excel_headers() -> None:
    from openpyxl import load_workbook
    import tempfile

    records = [price_compare.build_record(_scrape("1", price=480.0), _mapping("1", control=500.0, barcode="BC1"))]
    with tempfile.TemporaryDirectory() as d:
        path = write_excel(Path(d) / "x.xlsx", records)
        ws = load_workbook(path).active
        headers = [c.value for c in ws[1]]
    assert headers == [
        "商品ID", "商品标题名称", "条码", "聚水潭商品编码", "聚水潭商品名称",
        "淘系控价", "商品实时价格", "商品差价", "状态", "抓取时间", "截图路径",
    ]


def test_json_payload_fields() -> None:
    records = [price_compare.build_record(_scrape("1", price=480.0), _mapping("1", control=500.0, barcode="BC1", name="甲SKU"))]
    payload = build_json_payload(records, dry_run=False, source="page")
    assert payload["below_control_count"] == 1
    item = payload["below_control"][0]
    assert {"barcode", "jst_goods_code", "jst_goods_name", "taoxi_control_price", "diff_price"} <= set(item)
    assert item["taoxi_control_price"] == 500.0
