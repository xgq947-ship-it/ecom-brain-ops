from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.runtime import WorkflowRunner
from core.runtime.registry import discover_workflow
from core.task_registry import resolve_task

from workflows.tmcs_priority_promotion_plan_create import steps
from workflows.tmcs_priority_promotion_plan_create.workflow import build_workflow


def _make_promotion_file(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["猫超店铺推广SKU清单　（猫超）福安市启明工贸有限公司　2026-05"])
    worksheet.append([])
    worksheet.append(["🏆 优先推广（3 个）　全站推广（优先）→ 万相台关键词推广（叠加），目标推广占比 4%~8%"])
    worksheet.append(["店铺款式编码", "商品名称", "利润率", "毛利率", "销量(件)", "销售额(元)", "均价(元)", "退款率", "当前推广(元)", "推广占比"])
    worksheet.append(["1045053461141", "商品A", "40%", "50%", 15, 3000, 200, "10%", 0, "0%"])
    worksheet.append(["1048591828491", "商品B", "31%", "40%", 10, 2000, 200, "9%", 0, "0%"])
    worksheet.append(["766011668628", "商品C", "28%", "42%", 49, 18800, 384, "11%", 0, "0%"])
    worksheet.append([])
    worksheet.append(["✅ 次级推广（1 个）"])
    worksheet.append(["店铺款式编码", "商品名称"])
    worksheet.append(["SHOULD_NOT_USE", "次级商品"])
    workbook.save(path)
    workbook.close()
    return path


def _make_active_file(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["店铺款式编码"])
    worksheet.append(["1048591828491"])
    workbook.save(path)
    workbook.close()
    return path


def _make_master_file(path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品编码", "商品名称", "SKU编码", "条码", "货品编码"])
    worksheet.append(["1045053461141", "商品A", "SKU-A", "BAR-A", "HP-A"])
    worksheet.append(["1048591828491", "商品B", "SKU-B", "BAR-B", "HP-B"])
    worksheet.append(["766011668628", "商品C", "SKU-C", "BAR-C", "HP-C"])
    workbook.save(path)
    workbook.close()
    return path


def _runner(tmp_path: Path) -> WorkflowRunner:
    return WorkflowRunner(tmp_path / "runs")


def test_workflow_registers() -> None:
    workflow = discover_workflow("tmcs_priority_promotion_plan_create")
    assert workflow.id == "tmcs_priority_promotion_plan_create"
    assert [step.id for step in workflow.steps] == [
        "check_inputs",
        "load_priority_promotion_list",
        "load_active_promotion_list",
        "filter_not_active",
        "resolve_item_ids_for_plan",
        "build_create_plan_payloads",
        "create_zdx_plans",
        "sync_active_promotion_list",
        "write_outputs",
        "collect_outputs",
    ]


def test_chinese_entry_resolves() -> None:
    assert resolve_task("猫超优先推广自动建计划") == "猫超优先推广自动建计划"
    assert resolve_task("猫超智多星批量建计划") == "猫超优先推广自动建计划"


def test_month_defaults_to_last_month(monkeypatch) -> None:
    import datetime as real_datetime

    class FakeDate(real_datetime.date):
        @classmethod
        def today(cls):
            return real_datetime.date(2026, 6, 4)

    monkeypatch.setattr(steps, "date", FakeDate)
    assert steps._last_month() == "2026-05"


def test_dry_run_filters_active_and_builds_payloads(monkeypatch, tmp_path: Path) -> None:
    promotion = _make_promotion_file(tmp_path / "promotion.xlsx")
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    calls: list[tuple[str, list[str], bool]] = []

    def fake_run_child(workflow_id: str, args: list[str], *, dry_run: bool):
        calls.append((workflow_id, list(args), dry_run))
        raise AssertionError("dry-run 不应真实调用子 workflow")

    monkeypatch.setattr(steps, "_run_child_workflow", fake_run_child)

    run = _runner(tmp_path).run(
        build_workflow(),
        inputs={
            "dry_run": True,
            "args": [
                "--dry-run",
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
            ],
        },
        dry_run=True,
    )

    assert run.status == "dry_run_success"
    assert calls == []
    assert run.outputs["total_priority_count"] == 3
    assert run.outputs["already_active_count"] == 1
    assert run.outputs["to_create_count"] == 2
    assert run.outputs["created_count"] == 0
    assert run.outputs["failed_count"] == 0
    assert [item["store_style_code"] for item in run.outputs["to_create_items"]] == ["1045053461141", "766011668628"]
    assert all(item["target_workflow"] == "tmcs_zdx_fullsite_plan_create" for item in run.outputs["plan_payloads"])


def test_limit_applies_to_payloads(monkeypatch, tmp_path: Path) -> None:
    promotion = _make_promotion_file(tmp_path / "promotion.xlsx")
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    monkeypatch.setattr(steps, "_run_child_workflow", lambda *a, **k: (_ for _ in ()).throw(AssertionError("dry-run 不应调用子 workflow")))

    run = _runner(tmp_path).run(
        build_workflow(),
        inputs={
            "dry_run": True,
            "args": [
                "--dry-run",
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
                "--limit",
                "1",
            ],
        },
        dry_run=True,
    )
    assert run.status == "dry_run_success"
    assert len(run.outputs["plan_payloads"]) == 1


def test_missing_item_id_is_skipped(tmp_path: Path) -> None:
    promotion = _make_promotion_file(tmp_path / "promotion.xlsx")
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["商品编码", "商品名称", "SKU编码", "条码", "货品编码"])
    worksheet.append(["1045053461141", "商品A", "SKU-A", "BAR-A", "HP-A"])
    workbook.save(master)
    workbook.close()

    run = _runner(tmp_path).run(
        build_workflow(),
        inputs={
            "dry_run": True,
            "args": [
                "--dry-run",
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
            ],
        },
        dry_run=True,
    )
    assert run.status == "dry_run_success"
    skipped = run.outputs["skipped_items"]
    assert any(item["skip_reason"] == "missing_item_id" and item["store_style_code"] == "766011668628" for item in skipped)


def test_execute_marks_workflow_failed_when_any_child_plan_fails_after_sync(monkeypatch, tmp_path: Path) -> None:
    promotion = _make_promotion_file(tmp_path / "promotion.xlsx")
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    seen: list[tuple[str, list[str], bool]] = []

    class FakeRun:
        def __init__(self, status: str, outputs: dict, errors: list[str], run_id: str) -> None:
            self.status = status
            self.outputs = outputs
            self.errors = errors
            self.run_id = run_id

    def fake_run_child(workflow_id: str, args: list[str], *, dry_run: bool):
        seen.append((workflow_id, list(args), dry_run))
        item_id = args[args.index("--item-id") + 1]
        if item_id == "766011668628":
            return FakeRun("failed", {}, ["创建失败"], "run-child-fail")
        return FakeRun("success", {"created": True, "platform_plan_id": "PLAN-1"}, [], "run-child-ok")

    monkeypatch.setattr(steps, "_run_child_workflow", fake_run_child)

    run = _runner(tmp_path).run(
        build_workflow(),
        inputs={
            "args": [
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
                "--execute",
            ],
        },
        dry_run=False,
    )

    assert run.status == "failed"
    assert any("创建失败" in err for err in run.errors)
    assert len(seen) == 2
    assert all(call[0] == "tmcs_zdx_fullsite_plan_create" for call in seen)
    assert all("--execute" in call[1] for call in seen)
    assert all("--confirm-plan-name" in call[1] for call in seen)
    assert run.outputs["created_count"] == 1
    assert run.outputs["failed_count"] == 1
    assert run.outputs["active_promotion_appended_codes"] == ["1045053461141"]

    workbook = load_workbook(active, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        values = [row[0] for row in worksheet.iter_rows(values_only=True) if row and row[0]]
    finally:
        workbook.close()
    assert values == ["店铺款式编码", "1048591828491", "1045053461141"]


def test_auto_generate_source_calls_sales_workflow(monkeypatch, tmp_path: Path) -> None:
    promotion = tmp_path / "promotion.xlsx"
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    seen: list[tuple[str, list[str], bool]] = []

    class FakeRun:
        def __init__(self, status: str, outputs: dict, errors: list[str], run_id: str) -> None:
            self.status = status
            self.outputs = outputs
            self.errors = errors
            self.run_id = run_id

    def fake_run_child(workflow_id: str, args: list[str], *, dry_run: bool):
        seen.append((workflow_id, list(args), dry_run))
        _make_promotion_file(promotion)
        return FakeRun("dry_run_success", {}, [], "run-source")

    monkeypatch.setattr(steps, "_run_child_workflow", fake_run_child)

    run = _runner(tmp_path).run(
        build_workflow(),
        inputs={
            "dry_run": True,
            "args": [
                "--dry-run",
                "--auto-generate-source",
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
            ],
        },
        dry_run=True,
    )

    assert run.status == "dry_run_success"
    assert seen[0][0] == "jst_tmcs_shop_product_sales_analysis"
    assert "--output" in seen[0][1]
    assert run.outputs["generated_source_run_id"] == "run-source"


def test_output_file_written(monkeypatch, tmp_path: Path) -> None:
    promotion = _make_promotion_file(tmp_path / "promotion.xlsx")
    active = _make_active_file(tmp_path / "active.xlsx")
    master = _make_master_file(tmp_path / "master.xlsx")
    output = tmp_path / "result.json"
    monkeypatch.setattr(steps, "_run_child_workflow", lambda *a, **k: (_ for _ in ()).throw(AssertionError("不应调用子 workflow")))

    runner = _runner(tmp_path)
    run = runner.run(
        build_workflow(),
        inputs={
            "args": [
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
                "--output",
                str(output),
                "--execute",
            ],
        },
        dry_run=False,
    )

    assert run.status == "failed"
    assert not output.exists()

    monkeypatch.setattr(
        steps,
        "_run_child_workflow",
        lambda workflow_id, args, dry_run: type("FakeRun", (), {"status": "success", "outputs": {"created": True}, "errors": [], "run_id": "child"})(),
    )
    run = runner.run(
        build_workflow(),
        inputs={
            "args": [
                "--promotion-list-file",
                str(promotion),
                "--active-promotion-file",
                str(active),
                "--tmcs-master-file",
                str(master),
                "--daily-budget",
                "100",
                "--output",
                str(output),
                "--execute",
            ],
        },
        dry_run=False,
    )
    assert run.status == "success"
    assert output.exists()
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert "priority_items" in payload and "created_items" in payload
