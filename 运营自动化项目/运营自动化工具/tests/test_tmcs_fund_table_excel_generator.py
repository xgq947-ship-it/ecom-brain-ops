from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from workflows.tmcs_fund_table_generate.excel_generator import (
    FUND_TABLE_HEADERS,
    generate_fund_table,
    verify_fund_table,
)


def _png(path: Path, color: str) -> Path:
    image = Image.new("RGB", (240, 120), color=color)
    image.save(path)
    return path


def test_generate_new_fund_table_with_formulas_and_images(tmp_path: Path) -> None:
    receivable = _png(tmp_path / "receivable.png", "white")
    promotion = _png(tmp_path / "promotion.png", "blue")
    output = tmp_path / "猫超资金表_2026-05.xlsx"

    result = generate_fund_table(
        output_file=output,
        month="2026-05",
        receivable_amount=802.35,
        promotion_balance=600.0,
        reserve_balance=123.45,
        bank_card_balance=678.9,
        receivable_screenshot=receivable,
        promotion_screenshot=promotion,
    )

    assert result.output_file == output
    assert result.formula_check_result == {"Q2": True, "S2": True}

    workbook = load_workbook(output)
    sheet = workbook["店铺资金"]
    assert [sheet.cell(1, col).value for col in range(1, 20)] == FUND_TABLE_HEADERS
    assert sheet["A2"].value == "国清"
    assert sheet["B2"].value == "（猫超）福安市启明工贸有限公司（国清）"
    assert sheet["C2"].value == "12633507"
    assert sheet["D2"].value == "天猫超市"
    assert sheet["F2"].value == 50000
    assert sheet["I2"].value == 802.35
    assert sheet["J2"].value == 600.0
    assert sheet["M2"].value == 123.45
    assert sheet["N2"].value == 678.9
    assert sheet["Q2"].value == "=H2+I2+L2+M2+N2+O2+P2"
    assert sheet["S2"].value == "=Q2-R2"
    # 两张截图嵌入数据下一行同列：待收货款 → I3、推广账户余额 → J3
    assert len(sheet._images) == 2
    anchors = sorted((img.anchor._from.col, img.anchor._from.row) for img in sheet._images)
    assert anchors == [(8, 2), (9, 2)]  # I3=(col8,row2), J3=(col9,row2)，均为 0 基索引
    workbook.close()


def test_verify_fund_table_rejects_missing_formula(tmp_path: Path) -> None:
    receivable = _png(tmp_path / "receivable.png", "white")
    promotion = _png(tmp_path / "promotion.png", "blue")
    output = tmp_path / "bad.xlsx"
    generate_fund_table(
        output_file=output,
        month="2026-05",
        receivable_amount=1,
        promotion_balance=2,
        reserve_balance=0,
        bank_card_balance=0,
        receivable_screenshot=receivable,
        promotion_screenshot=promotion,
    )
    workbook = load_workbook(output)
    sheet = workbook["店铺资金"]
    sheet["Q2"] = 3
    workbook.save(output)
    workbook.close()

    result = verify_fund_table(output)

    assert result["Q2"] is False
    assert result["S2"] is True
