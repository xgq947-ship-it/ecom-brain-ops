from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.runtime import WorkflowRunner
from core.runtime.registry import discover_workflow
from core.task_registry import resolve_task
from tasks.tmcs_product_code_lookup import main as task_entry
from workflows.tmcs_product_code_lookup.workflow import build_workflow


HEADERS = ["商品编码", "商品名称", "商品上下架状态", "SKU编码", "条码", "淘系品牌名称"]


def _make_source(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["7620001", "奥克斯按摩靠垫颈椎按摩器", "上架", "SKU001", "AUAMKDK1606", "AUX/奥克斯"])
    sheet.append(["7620002", "奥克斯按摩椅家用全身", "上架", "SKU002", "AUAMKDK1607", "AUX/奥克斯"])
    sheet.append(["7620003", "美的电风扇落地扇", "上架", "SKU003", "MDFAN2024", "Midea/美的"])
    sheet.append(["7620004", "奥克斯下架款按摩仪", "下架", "SKU004", "AUAMKDK9999", "AUX/奥克斯"])
    workbook.save(path)
    workbook.close()
    return path


def _run(tmp_path: Path, args: list[str], *, dry_run: bool = True):
    runner = WorkflowRunner(tmp_path / "runs")
    run = runner.run(build_workflow(), inputs={"dry_run": dry_run, "args": args}, dry_run=dry_run)
    return runner, run


def _collect_outputs(runner: WorkflowRunner) -> dict:
    return json.loads((runner.last_run_dir / "steps" / "collect_outputs.json").read_text(encoding="utf-8"))["outputs"]


def test_workflow_registers() -> None:
    workflow = discover_workflow("tmcs_product_code_lookup")
    assert workflow.id == "tmcs_product_code_lookup"
    assert [s.id for s in workflow.steps] == [
        "check_inputs",
        "load_tmcs_products",
        "fuzzy_match_products",
        "collect_outputs",
    ]


def test_chinese_entry_resolves() -> None:
    for alias in ("猫超商品编码查询", "查询猫超商品编码", "型号查商品编码", "猫超型号查询"):
        assert resolve_task(alias) == "tmcs_product_code_lookup"


def test_main_routes_to_workflow(monkeypatch) -> None:
    calls: list[list[str]] = []
    monkeypatch.setattr(sys, "argv", ["tmcs_product_code_lookup", "--model", "AUX", "--dry-run"])
    monkeypatch.setattr(task_entry, "_run_workflow", lambda args: calls.append(list(args)) or 0, raising=False)
    assert task_entry.main() == 0
    assert calls == [["tmcs_product_code_lookup", "--model", "AUX", "--dry-run"]]


def test_model_is_required(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    runner, run = _run(tmp_path, ["--dry-run", "--source-file", str(source)])
    assert run.status == "failed"
    assert any("--model" in error for error in run.errors)


def test_only_online_products_matched(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    runner, run = _run(tmp_path, ["--dry-run", "--source-file", str(source), "--model", "AUAMKDK", "--min-score", "0.5"])
    assert run.status == "dry_run_success"
    outputs = _collect_outputs(runner)
    codes = {item["product_code"] for item in outputs["results"]}
    assert "7620004" not in codes  # 下架款不出现


def test_brand_filter(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    runner, run = _run(
        tmp_path,
        ["--dry-run", "--source-file", str(source), "--model", "电风扇", "--brand", "美的", "--min-score", "0.3"],
    )
    assert run.status == "dry_run_success"
    outputs = _collect_outputs(runner)
    assert outputs["matched_count"] >= 1
    assert all(item["brand"] == "Midea/美的" for item in outputs["results"])


def test_fuzzy_match_and_sorted(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    runner, run = _run(
        tmp_path,
        ["--dry-run", "--source-file", str(source), "--model", "AUAMKDK1606", "--min-score", "0.5"],
    )
    outputs = _collect_outputs(runner)
    assert outputs["results"]
    assert outputs["results"][0]["barcode"] == "AUAMKDK1606"
    assert outputs["results"][0]["product_code"] == "7620001"
    scores = [item["match_score"] for item in outputs["results"]]
    assert scores == sorted(scores, reverse=True)


def test_missing_field_reports_clear_error(tmp_path: Path) -> None:
    bad = tmp_path / "bad.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品编码", "商品名称", "商品上下架状态", "SKU编码"])  # 缺条码
    sheet.append(["A", "测试", "上架", "S"])
    workbook.save(bad)
    workbook.close()
    runner, run = _run(tmp_path, ["--dry-run", "--source-file", str(bad), "--model", "测试"])
    assert run.status == "failed"
    assert any("条码" in error and "缺少必需字段" in error for error in run.errors)


def _make_source_dup_sku(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_A", "SUH502", "SUPOR/苏泊尔"])
    sheet.append(["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_B", "SUH502", "SUPOR/苏泊尔"])
    workbook.save(path)
    workbook.close()
    return path


def test_default_dedupes_by_product_code(tmp_path: Path) -> None:
    source = _make_source_dup_sku(tmp_path / "src.xlsx")
    runner, run = _run(tmp_path, ["--dry-run", "--source-file", str(source), "--model", "SUH502", "--min-score", "0.5"])
    assert run.status == "dry_run_success"
    outputs = _collect_outputs(runner)
    assert outputs["dedupe_by"] == "product_code"
    assert outputs["matched_count"] == 1
    assert outputs["results"][0]["product_code"] == "SPU100"


def test_by_sku_flag_keeps_each_sku(tmp_path: Path) -> None:
    source = _make_source_dup_sku(tmp_path / "src.xlsx")
    runner, run = _run(
        tmp_path,
        ["--dry-run", "--source-file", str(source), "--model", "SUH502", "--min-score", "0.5", "--by-sku"],
    )
    outputs = _collect_outputs(runner)
    assert outputs["dedupe_by"] == "sku"
    assert outputs["matched_count"] == 2
    assert {r["sku_code"] for r in outputs["results"]} == {"SKU_A", "SKU_B"}


def test_no_match_returns_zero(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    runner, run = _run(
        tmp_path,
        ["--dry-run", "--source-file", str(source), "--model", "绝不存在XYZ型号", "--min-score", "0.9"],
    )
    assert run.status == "dry_run_success"
    outputs = _collect_outputs(runner)
    assert outputs["matched_count"] == 0
    assert outputs["results"] == []
    assert "message" in outputs


def test_dry_run_does_not_write_output(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    out = tmp_path / "result.json"
    runner, run = _run(
        tmp_path,
        ["--dry-run", "--source-file", str(source), "--model", "AUAMKDK1606", "--output", str(out)],
    )
    assert run.status == "dry_run_success"
    assert not out.exists()


def test_execute_writes_output_and_records_artifact(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    out = tmp_path / "result.json"
    runner, run = _run(
        tmp_path,
        ["--source-file", str(source), "--model", "AUAMKDK1606", "--output", str(out)],
        dry_run=False,
    )
    assert run.status == "success"
    assert out.exists()
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert payload["matched_count"] >= 1
    assert payload["query"]["model"] == "AUAMKDK1606"
    assert any(a.role == "output" and a.path == str(out) for a in run.artifacts)


def test_source_excel_not_modified(tmp_path: Path) -> None:
    source = _make_source(tmp_path / "src.xlsx")
    before = hashlib.sha256(source.read_bytes()).hexdigest()
    _run(tmp_path, ["--dry-run", "--source-file", str(source), "--model", "AUAMKDK1606"])
    after = hashlib.sha256(source.read_bytes()).hexdigest()
    assert before == after
