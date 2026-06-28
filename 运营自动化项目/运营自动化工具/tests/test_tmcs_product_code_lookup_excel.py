from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from workflows.tmcs_product_code_lookup.excel_lookup import (
    FieldResolutionError,
    fuzzy_match_products,
    load_online_products,
    resolve_fields,
)


HEADERS = [
    "商品编码",
    "商品名称",
    "商品上下架状态",
    "SKU编码",
    "条码",
    "淘系品牌名称",
]


def _make_file(path: Path, rows: list[list]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _sample_rows() -> list[list]:
    return [
        # product_code, name, status, sku_code, barcode, brand
        ["7620001", "奥克斯按摩靠垫颈椎按摩器", "上架", "SKU001", "AUAMKDK1606", "AUX/奥克斯"],
        ["7620002", "奥克斯按摩椅家用全身", "上架", "SKU002", "AUAMKDK1607", "AUX/奥克斯"],
        ["7620003", "美的电风扇落地扇", "上架", "SKU003", "MDFAN2024", "Midea/美的"],
        ["7620004", "奥克斯下架款按摩仪", "下架", "SKU004", "AUAMKDK9999", "AUX/奥克斯"],
    ]


def test_load_filters_online_only(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    products, resolved = load_online_products(path)
    assert {"status", "product_code", "sku_code", "barcode", "product_name", "brand"} <= set(resolved)
    # 只保留 3 条上架商品（下架那条被过滤）
    assert len(products) == 3
    assert all(p["product_code"] != "7620004" for p in products)


def test_missing_required_field_raises_clear_error(tmp_path: Path):
    # 缺少「条码」必需字段
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["商品编码", "商品名称", "商品上下架状态", "SKU编码"])
    sheet.append(["A", "测试", "上架", "S"])
    path = tmp_path / "bad.xlsx"
    workbook.save(path)
    workbook.close()
    with pytest.raises(FieldResolutionError) as exc:
        load_online_products(path)
    assert "条码" in str(exc.value)


def test_resolve_fields_uses_candidate_headers():
    headers = {"猫超商品编码": 0, "标题": 1, "状态": 2, "平台SKUID": 3, "barcode": 4}
    resolved, missing = resolve_fields(headers)
    assert resolved["product_code"] == 0
    assert resolved["product_name"] == 1
    assert resolved["status"] == 2
    assert resolved["sku_code"] == 3
    assert resolved["barcode"] == 4
    assert missing == []


def test_model_fuzzy_match_hits(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="AUAMKDK1606", min_score=0.5)
    assert results
    assert results[0]["barcode"] == "AUAMKDK1606"
    assert results[0]["product_code"] == "7620001"
    assert results[0]["match_score"] >= 0.9


def test_results_sorted_by_score_desc(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="按摩", min_score=0.3, limit=10)
    scores = [r["match_score"] for r in results]
    assert scores == sorted(scores, reverse=True)


def test_brand_filter_applies(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="按摩", brand="美的", min_score=0.3, limit=10)
    # 「美的」品牌里没有「按摩」字样，应当被品牌过滤掉
    assert results == []
    results2 = fuzzy_match_products(products, model="电风扇", brand="美的", min_score=0.3, limit=10)
    assert results2 and all(r["brand"] == "Midea/美的" for r in results2)


def test_dedupe_by_product_code_by_default(tmp_path: Path):
    # 同一商品编码挂两个 SKU，默认按商品编码去重后只保留一条
    rows = [
        ["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_A", "SUH502", "SUPOR/苏泊尔"],
        ["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_B", "SUH502", "SUPOR/苏泊尔"],
    ]
    path = _make_file(tmp_path / "src.xlsx", rows)
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="SUH502", min_score=0.5)
    assert len(results) == 1
    assert results[0]["product_code"] == "SPU100"


def test_by_sku_keeps_each_sku(tmp_path: Path):
    rows = [
        ["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_A", "SUH502", "SUPOR/苏泊尔"],
        ["SPU100", "苏泊尔足疗机H5全自动", "上架", "SKU_B", "SUH502", "SUPOR/苏泊尔"],
    ]
    path = _make_file(tmp_path / "src.xlsx", rows)
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="SUH502", min_score=0.5, dedupe_by="sku")
    assert len(results) == 2
    assert {r["sku_code"] for r in results} == {"SKU_A", "SKU_B"}


def test_no_match_returns_empty(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    products, _ = load_online_products(path)
    results = fuzzy_match_products(products, model="完全不存在的型号XYZ", min_score=0.9)
    assert results == []


def test_does_not_modify_source_excel(tmp_path: Path):
    path = _make_file(tmp_path / "src.xlsx", _sample_rows())
    before = hashlib.sha256(path.read_bytes()).hexdigest()
    load_online_products(path)
    fuzzy_match_products(load_online_products(path)[0], model="按摩", min_score=0.3)
    after = hashlib.sha256(path.read_bytes()).hexdigest()
    assert before == after
    # 文件仍可正常打开
    wb = load_workbook(path, read_only=True)
    wb.close()
