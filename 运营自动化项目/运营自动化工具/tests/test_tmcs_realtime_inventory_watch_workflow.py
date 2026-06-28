"""猫超库存实时监测 workflow 集成测试。"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from core.runtime import WorkflowRunner
from core.runtime.registry import discover_workflow
from core.task_registry import resolve_task

from workflows.tmcs_realtime_inventory_watch import steps
from workflows.tmcs_realtime_inventory_watch.workflow import build_workflow


# ── 测试 Excel 工厂 ──────────────────────────────────────────────────────────
def _xlsx(path: Path, headers: list[str], rows: list[list]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def _goods_file(path: Path) -> Path:
    return _xlsx(
        path,
        ["SKU编码", "条码", "商品上下架状态"],
        [
            ["SKU_RISK", "BC_RISK", "上架"],
            ["SKU_SAFE", "BC_SAFE", "上架"],
            ["SKU_OFF", "BC_OFF", "下架"],  # 下架应被过滤
        ],
    )


def _jst_file(path: Path) -> Path:
    return _xlsx(
        path,
        ["商品编码", "实际库存数", "订单占有数", "品牌"],
        [
            ["BC_RISK", 8, 3, "苏泊尔"],   # 剩余 5 (<20)
            ["BC_SAFE", 50, 0, "奥克斯"],  # 剩余 50 (>=20)
            ["BC_OFF", 1, 0, "苏泊尔"],
            ["BC_OTHER", 1, 0, "美的"],    # 非目标品牌
        ],
    )


def _tmcs_file(path: Path) -> Path:
    return _xlsx(
        path,
        ["平台SKUID", "专享现货库存可售量", "共享现货库存可售量", "商家仓code"],
        [
            ["SKU_RISK", 10, 5, "mc_aokesi_suolong"],  # 可售 15 > 剩余 5 -> 风险
            ["SKU_SAFE", 99, 0, "mc_aokesi_suolong"],
            ["SKU_RISK", 1, 1, "other_wh"],            # 其他仓应被过滤
        ],
    )


def _run(tmp_path: Path, extra_args: list[str], *, dry_run: bool):
    runner = WorkflowRunner(tmp_path / "runs")
    args = list(extra_args)
    if dry_run and "--dry-run" not in args:
        args.append("--dry-run")
    run = runner.run(build_workflow(), inputs={"dry_run": dry_run, "args": args}, dry_run=dry_run)
    return run, runner


def _step_outputs(runner: WorkflowRunner, step_id: str) -> dict:
    path = runner.last_run_dir / "steps" / f"{step_id}.json"
    return json.loads(path.read_text(encoding="utf-8"))["outputs"]


def _base_args(tmp_path: Path) -> list[str]:
    goods = _goods_file(tmp_path / "goods.xlsx")
    jst = _jst_file(tmp_path / "jst.xlsx")
    tmcs = _tmcs_file(tmp_path / "tmcs.xlsx")
    return [
        "--maochao-goods-file", str(goods),
        "--use-local-jst-file", str(jst),
        "--use-local-tmcs-stock-file", str(tmcs),
    ]


# 1. workflow 可以注册
def test_workflow_registers():
    wf = discover_workflow("tmcs_realtime_inventory_watch")
    assert wf.id == "tmcs_realtime_inventory_watch"
    assert [s.id for s in wf.steps] == [
        "check_inputs",
        "refresh_jst_product_data",
        "refresh_tmcs_stock_data",
        "load_maochao_goods",
        "load_jst_products",
        "load_tmcs_stock",
        "build_inventory_table",
        "detect_inventory_risks",
        "write_outputs",
        "notify_if_needed",
        "collect_outputs",
    ]


# 2. 中文入口可以解析
def test_chinese_alias_resolves():
    assert resolve_task("猫超库存实时监测") == "tmcs_realtime_inventory_watch"
    assert resolve_task("库存实时监测") == "tmcs_realtime_inventory_watch"
    assert resolve_task("猫超库存风险监测") == "tmcs_realtime_inventory_watch"
    assert resolve_task("聚水潭猫超库存对比") == "tmcs_realtime_inventory_watch"
    assert resolve_task("猫超可售库存预警") == "tmcs_realtime_inventory_watch"


def test_default_maochao_goods_file_uses_configured_path(monkeypatch, tmp_path: Path) -> None:
    configured = tmp_path / "猫超商品列表导出 (最新）.xlsx"

    def fake_get_path(name: str) -> Path:
        if name != "tmall_goods_master_file":
            raise KeyError(name)
        return configured

    monkeypatch.setattr(steps, "get_path", fake_get_path)

    flags = steps._parse_flags(SimpleNamespace(inputs={"args": []}, dry_run=True))

    assert flags.maochao_goods_file == str(configured)


# 3/4/5. 读取并筛选三张表
def test_loads_and_filters_three_tables(tmp_path):
    run, runner = _run(tmp_path, _base_args(tmp_path), dry_run=True)
    assert run.status in {"success", "dry_run_success"}
    assert _step_outputs(runner, "load_maochao_goods")["active_tmcs_goods_rows"] == 2  # 下架过滤
    assert _step_outputs(runner, "load_jst_products")["jst_rows"] == 3  # 美的过滤
    assert _step_outputs(runner, "load_tmcs_stock")["tmcs_stock_rows"] == 2  # 其他仓过滤


# 6/7. 表4 关联与剩余库存
def test_inventory_table_join_and_remaining(tmp_path):
    _, runner = _run(tmp_path, _base_args(tmp_path), dry_run=True)
    assert _step_outputs(runner, "build_inventory_table")["matched_rows"] == 2


# 8/9. 风险输出（剩余<20 且 猫超可售>剩余）
def test_risk_detected(tmp_path):
    _, runner = _run(tmp_path, _base_args(tmp_path), dry_run=True)
    out = _step_outputs(runner, "detect_inventory_risks")
    assert out["risk_count"] == 1
    assert out["low_stock_count"] == 1
    risk = out["risk_items"][0]
    assert risk["sku_code"] == "SKU_RISK"
    assert risk["actual_stock"] == 8  # 聚水潭实际库存(实际库存数)
    assert risk["tmcs_total_sellable_stock"] == 15


# 10. 聚水潭实际库存 >= threshold 不记录（阈值调到 5 时 RISK 的实际库存 8 不再 < 5）
def test_no_record_when_actual_ge_threshold(tmp_path):
    args = _base_args(tmp_path) + ["--threshold", "5"]
    _, runner = _run(tmp_path, args, dry_run=True)
    assert _step_outputs(runner, "detect_inventory_risks")["risk_count"] == 0


# 11. 不剔除「猫超可售=0/偏低」：实际库存<20 且能关联猫超库存即记录
def test_record_kept_even_if_tmcs_sellable_low(tmp_path):
    goods = _goods_file(tmp_path / "goods.xlsx")
    jst = _jst_file(tmp_path / "jst.xlsx")
    tmcs = _xlsx(
        tmp_path / "tmcs.xlsx",
        ["平台SKUID", "专享现货库存可售量", "共享现货库存可售量", "商家仓code"],
        [["SKU_RISK", 0, 0, "mc_aokesi_suolong"]],  # 猫超可售 0，仍记录
    )
    args = [
        "--maochao-goods-file", str(goods),
        "--use-local-jst-file", str(jst),
        "--use-local-tmcs-stock-file", str(tmcs),
    ]
    _, runner = _run(tmp_path, args, dry_run=True)
    out = _step_outputs(runner, "detect_inventory_risks")
    assert out["risk_count"] == 1
    assert out["risk_items"][0]["tmcs_total_sellable_stock"] == 0


# 12. 缺字段时 workflow 失败并给清晰错误
def test_missing_field_fails(tmp_path):
    bad_jst = _xlsx(tmp_path / "bad.xlsx", ["商品编码", "品牌"], [["BC_RISK", "苏泊尔"]])
    goods = _goods_file(tmp_path / "goods.xlsx")
    tmcs = _tmcs_file(tmp_path / "tmcs.xlsx")
    args = [
        "--maochao-goods-file", str(goods),
        "--use-local-jst-file", str(bad_jst),
        "--use-local-tmcs-stock-file", str(tmcs),
    ]
    run, runner = _run(tmp_path, args, dry_run=True)
    assert run.status == "failed"
    out = _step_outputs(runner, "load_jst_products")
    # step 失败，错误信息含字段名
    step_json = json.loads((runner.last_run_dir / "steps" / "load_jst_products.json").read_text(encoding="utf-8"))
    assert any("实际库存" in e for e in step_json.get("errors", [])) or "实际库存" in json.dumps(out, ensure_ascii=False)


# 13. dry-run + 本地文件 不触发真实下载
def test_dry_run_local_files_no_download(tmp_path, monkeypatch):
    called: list = []

    def boom(command, interactive_recovery=None):
        called.append(command)
        raise AssertionError("dry-run/本地文件不应调用平台下载")

    monkeypatch.setattr(steps, "run_ops_json", boom)
    run, _ = _run(tmp_path, _base_args(tmp_path), dry_run=True)
    assert run.status in {"success", "dry_run_success"}
    assert called == []


# 13b. 真实运行（非 dry-run）默认实时下载两份源数据，不复用旧文件
def test_real_run_downloads_fresh_by_default(tmp_path, monkeypatch):
    goods = _goods_file(tmp_path / "goods.xlsx")
    fresh_jst = _jst_file(tmp_path / "fresh_jst.xlsx")
    fresh_tmcs = _tmcs_file(tmp_path / "fresh_tmcs.xlsx")
    calls: list = []

    def fake_run_ops_json(command, interactive_recovery=None):
        calls.append(list(command))
        if "jst" in command and "product" in command:
            return {"success": True, "data": {"output_path": str(fresh_jst), "downloaded": True}}
        if "tmcs" in command and "inventory" in command:
            return {"success": True, "data": {"output_path": str(fresh_tmcs)}}
        raise AssertionError(f"未预期的平台调用：{command}")

    monkeypatch.setattr(steps, "run_ops_json", fake_run_ops_json)
    # 真实运行：不传 --use-local-*，应分别实时下载 jst / tmcs
    run, runner = _run(tmp_path, ["--maochao-goods-file", str(goods)], dry_run=False)
    assert run.status == "success"
    assert any("jst" in c and "product" in c and "sync" in c for c in calls)
    assert any("tmcs" in c and "inventory" in c and "export" in c for c in calls)
    assert _step_outputs(runner, "refresh_jst_product_data")["fresh"] is True
    assert _step_outputs(runner, "refresh_tmcs_stock_data")["fresh"] is True
    assert _step_outputs(runner, "detect_inventory_risks")["risk_count"] == 1


# 13c. --output 生成的 xlsx 含主表+子表两个 sheet，子表为聚水潭≥20且猫超可售<50
def test_output_has_low_tmcs_subsheet(tmp_path):
    import openpyxl

    goods = _xlsx(
        tmp_path / "goods.xlsx",
        ["SKU编码", "条码", "商品上下架状态", "商品名称"],
        [
            ["SKU_RISK", "BC_RISK", "上架", "风险品"],   # 聚水潭实际 8 -> 主表
            ["SKU_SUB", "BC_SUB", "上架", "子表品"],      # 聚水潭实际 30, 猫超可售 30 -> 子表
        ],
    )
    jst = _xlsx(
        tmp_path / "jst.xlsx",
        ["商品编码", "实际库存数", "订单占有数", "品牌"],
        [["BC_RISK", 8, 0, "苏泊尔"], ["BC_SUB", 30, 0, "奥克斯"]],
    )
    tmcs = _xlsx(
        tmp_path / "tmcs.xlsx",
        ["平台SKUID", "专享现货库存可售量", "共享现货库存可售量", "商家仓code"],
        [["SKU_RISK", 10, 5, "mc_aokesi_suolong"], ["SKU_SUB", 20, 10, "mc_aokesi_suolong"]],
    )
    output = tmp_path / "out.xlsx"
    args = [
        "--maochao-goods-file", str(goods),
        "--use-local-jst-file", str(jst),
        "--use-local-tmcs-stock-file", str(tmcs),
        "--output", str(output),
    ]
    _, runner = _run(tmp_path, args, dry_run=True)
    out = _step_outputs(runner, "detect_inventory_risks")
    assert out["risk_count"] == 1 and out["low_tmcs_count"] == 1
    wb = openpyxl.load_workbook(output)
    assert wb.sheetnames == ["库存风险", "猫超低库存"]
    main = list(wb["库存风险"].iter_rows(values_only=True))
    sub = list(wb["猫超低库存"].iter_rows(values_only=True))
    assert main[0] == ("SKU编码", "商品名称", "聚水潭实际库存", "猫超实际库存")
    assert main[1][0] == "SKU_RISK"
    assert sub[1][0] == "SKU_SUB" and sub[1][2] == 30 and sub[1][3] == 30


# 14. --output 生成结果文件并记录 Artifact
def test_output_file_and_artifact(tmp_path):
    output = tmp_path / "risk.xlsx"
    args = _base_args(tmp_path) + ["--output", str(output)]
    run, runner = _run(tmp_path, args, dry_run=True)
    assert output.exists()
    out = _step_outputs(runner, "write_outputs")
    assert out["output_written"] is True
    artifacts = json.loads((runner.last_run_dir / "artifacts.json").read_text(encoding="utf-8"))
    assert any(a["role"] == "output" and a["path"] == str(output) for a in artifacts)
