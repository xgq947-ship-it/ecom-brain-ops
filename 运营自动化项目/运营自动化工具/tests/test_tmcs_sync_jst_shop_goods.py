from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_DIR = Path(__file__).resolve().parents[1] / "workflows" / "tmcs_sync_jst_shop_goods"
PROJECT_ROOT = Path(__file__).resolve().parents[1]

from workflows.tmcs_sync_jst_shop_goods.excel_builder import (  # noqa: E402
    CORRESPONDENCE_HEADERS,
    IMPORT_HEADERS,
    build_correspondence_rows,
    build_correspondence_workbook,
    build_import_workbooks,
    build_rows,
    load_jst_code_pool,
    match_corresponding_code,
)
from workflows.tmcs_sync_jst_shop_goods.input_loader import load_item_ids_from_excel, parse_item_ids  # noqa: E402
from workflows.tmcs_sync_jst_shop_goods import cli_client as skill_cli_client  # noqa: E402

from core.task_registry import resolve_task, task_scripts  # noqa: E402
import tasks.tmcs_sync_jst_shop_goods.main as task_entry  # noqa: E402


def test_skill_does_not_contain_platform_browser_automation_code() -> None:
    forbidden = ["playwright", "connect_over_cdp", "cookie", "localStorage", "sessionStorage", "http://", "https://", "selector"]
    for path in MODULE_DIR.glob("*.py"):
        if path.name == "cli_client.py":
            continue
        text = path.read_text(encoding="utf-8")
        lowered = text.lower()
        for token in forbidden:
            assert token.lower() not in lowered, f"{path.name} should not contain platform-side token {token}"


def test_parse_item_ids_dedupes_and_preserves_order() -> None:
    assert parse_item_ids("123, 234,123,,345") == ["123", "234", "345"]


def test_load_item_ids_from_excel_accepts_alias_headers(tmp_path: Path) -> None:
    path = tmp_path / "商品ID列表.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["商品ID", "备注"])
    ws.append(["1052534376394", "a"])
    ws.append(["1052534376394", "duplicate"])
    ws.append(["6247519890566", "b"])
    wb.save(path)

    assert load_item_ids_from_excel(path) == ["1052534376394", "6247519890566"]


def test_build_rows_maps_stock_to_jst_import_shape() -> None:
    import_rows, failures = build_rows(
        requested_item_ids=["1052534376394"],
        stock_rows=[
            {
                "platform_item_id": "1052534376394",
                "platform_sku_id": "6247519890565",
                "supplier_goods_id": "SUP-001",
                "merchant_goods_code": "MGC-001",
            }
        ],
    )

    assert failures == []
    assert import_rows == [
        {
            "线上款式编码": "1052534376394",
            "线上商品编码": "MGC-001",
            "线上国标码": "",
            "平台店铺款式编码": "1052534376394",
            "平台店铺商品编码": "SUP-001",
            "原始商品编码": "MGC-001",
            "线上商品名称": "",
            "线上颜色规格": "",
            "商品标识": "Retail",
        }
    ]


def test_build_import_workbooks_writes_text_cells(tmp_path: Path) -> None:
    result = build_import_workbooks(
        import_rows=[
            {
                "线上款式编码": "1052534376394",
                "线上商品编码": "MGC-001",
                "线上国标码": "",
                "平台店铺款式编码": "1052534376394",
                "平台店铺商品编码": "SUP-001",
                "原始商品编码": "MGC-001",
                "线上商品名称": "",
                "线上颜色规格": "",
                "商品标识": "Retail",
            }
        ],
        failures=[],
        output_dir=tmp_path,
        timestamp="20260518_120000",
    )

    wb = load_workbook(result["import_path"])
    ws = wb.active
    assert [cell.value for cell in ws[1]] == IMPORT_HEADERS
    assert ws["A2"].value == "1052534376394"
    assert ws["A2"].number_format == "@"


def test_skill_real_platform_call_uses_shared_interactive_recovery(monkeypatch) -> None:
    observed: dict[str, object] = {}

    def fake_run_ops_json(args, **kwargs):
        observed["args"] = args
        observed["kwargs"] = kwargs
        return {"success": True, "data": {"rows": []}}

    monkeypatch.setattr(skill_cli_client, "run_ops_json", fake_run_ops_json)

    assert skill_cli_client.query_tmcs_stock(item_ids=["1001"], warehouse_code="WH") == []
    assert observed["kwargs"] == {"interactive_recovery": True}


def test_formal_task_entry_resolves_chinese_triggers() -> None:
    assert resolve_task("聚水潭商品信息同步猫超") == "tmcs_sync_jst_shop_goods"
    assert resolve_task("猫超商品信息同步聚水潭") == "tmcs_sync_jst_shop_goods"
    assert resolve_task("平台商品ID同步聚水潭") == "tmcs_sync_jst_shop_goods"
    assert task_scripts()["tmcs_sync_jst_shop_goods"] == PROJECT_ROOT / "tasks" / "tmcs_sync_jst_shop_goods" / "main.py"


def test_pickup_watch_resolves_as_business_task() -> None:
    assert resolve_task("聚水潭揽收监控") == "jst_pickup_watch"
    assert task_scripts()["jst_pickup_watch"] == PROJECT_ROOT / "tasks" / "jst_pickup_watch.py"


def test_formal_task_entry_routes_to_workflow_without_skill_runpy(monkeypatch) -> None:
    calls: list[list[str]] = []
    monkeypatch.setattr(sys, "argv", ["main.py", "--item-ids", "1052305450766", "--dry-run"])
    monkeypatch.setattr(task_entry, "_run_workflow", lambda args: calls.append(list(args)) or 0, raising=False)

    assert task_entry.main() == 0
    assert not hasattr(task_entry, "runpy")
    assert calls == [["tmcs_sync_jst_shop_goods", "--item-ids", "1052305450766", "--dry-run"]]


def _make_jst_master(path: Path, codes: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["款式编码", "商品编码", "商品名称"])
    for code in codes:
        ws.append(["KS", code, "名"])
    wb.save(path)


def test_match_corresponding_code_exact_fuzzy_and_miss() -> None:
    exact_map = {"AOK-001": "AOK-001", "AOK-002A": "AOK-002A"}
    normalized_codes = ["AOK-001", "AOK-002A"]
    # 精确匹配
    assert match_corresponding_code("aok-001", exact_map, normalized_codes) == "AOK-001"
    # 尾部 .0 去除后精确匹配
    assert match_corresponding_code("AOK-001.0", exact_map, normalized_codes) == "AOK-001"
    # 唯一模糊匹配（线上码是池内某码的子串/超串）
    assert match_corresponding_code("AOK-002", exact_map, normalized_codes) == "AOK-002A"
    # 未命中：保留归一化原值
    assert match_corresponding_code("xyz-999", exact_map, normalized_codes) == "XYZ-999"
    # 空值
    assert match_corresponding_code("", exact_map, normalized_codes) == ""


def test_match_corresponding_code_multiple_keeps_original() -> None:
    exact_map = {"AOK-1": "AOK-1", "AOK-12": "AOK-12"}
    normalized_codes = ["AOK-1", "AOK-12"]
    # "AOK-1" 同时是 "AOK-1"(精确) -> 命中精确，先走精确
    assert match_corresponding_code("AOK-1", exact_map, normalized_codes) == "AOK-1"
    # "AOK" 是两者子串 -> 多义，保留归一化原值
    assert match_corresponding_code("AOK", exact_map, normalized_codes) == "AOK"


def test_build_correspondence_rows_maps_template_columns() -> None:
    import_rows = [
        {
            "线上款式编码": "1055840134894",
            "线上商品编码": "AOK-001",
            "平台店铺款式编码": "1055840134894",
            "平台店铺商品编码": "SUP-1",
            "线上商品名称": "",
            "线上颜色规格": "",
        }
    ]
    rows = build_correspondence_rows(import_rows, {"AOK-001": "AOK-001"}, ["AOK-001"])
    assert list(rows[0].keys()) == CORRESPONDENCE_HEADERS
    assert rows[0]["线上商品编码"] == "AOK-001"
    assert rows[0]["对应商品编码"] == "AOK-001"
    assert rows[0]["平台店铺商品编码"] == "SUP-1"


def test_build_correspondence_workbook_writes_file(tmp_path: Path) -> None:
    master = tmp_path / "聚水潭商品资料.xlsx"
    _make_jst_master(master, ["AOK-001", "AOK-002A"])
    import_rows = [
        {"线上款式编码": "I1", "线上商品编码": "AOK-001", "平台店铺款式编码": "I1", "平台店铺商品编码": "S1", "线上商品名称": "", "线上颜色规格": ""},
        {"线上款式编码": "I2", "线上商品编码": "AOK-002", "平台店铺款式编码": "I2", "平台店铺商品编码": "S2", "线上商品名称": "", "线上颜色规格": ""},
    ]
    result = build_correspondence_workbook(import_rows=import_rows, master_path=master, output_dir=tmp_path)
    out = Path(result["correspondence_path"])
    assert out.name == "猫超商品对应关系导入表.xlsx"
    assert out.exists()
    assert result["correspondence_rows"] == 2
    assert result["matched_rows"] == 1  # AOK-002 模糊匹配到 AOK-002A 算"匹配成功改写"
    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == CORRESPONDENCE_HEADERS
    data = [tuple(c.value for c in r) for r in ws.iter_rows(min_row=2, values_only=False)]
    # AOK-002 应被改写为 AOK-002A
    assert data[1][-1] == "AOK-002A"


def test_load_jst_code_pool_requires_product_code_column(tmp_path: Path) -> None:
    bad = tmp_path / "bad.xlsx"
    wb = Workbook()
    wb.active.append(["款式编码", "名称"])
    wb.save(bad)
    try:
        load_jst_code_pool(bad)
    except ValueError as exc:
        assert "商品编码" in str(exc)
    else:
        raise AssertionError("应因缺少商品编码列报错")
