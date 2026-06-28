from pathlib import Path
import sys
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from workflows.buyer_show import packager as buyer_show
from tasks import buyer_show as task_entry
from workflows.buyer_show.packager import bucket_assignments_by_brusher, date_suffix_for_filename, grouped_sources, plan_group_batches, verify_group_image_counts


class BuyerShowTests(unittest.TestCase):
    def test_bucket_assignments_by_brusher_splits_and_preserves_order(self):
        assignments = [
            ({"order_id": "1", "brusher": "唐杨"}, ("g1", [Path("a.jpg")])),
            ({"order_id": "2", "brusher": "小蝴蝶"}, ("g2", [Path("b.jpg")])),
            ({"order_id": "3", "brusher": "唐杨"}, ("g3", [Path("c.jpg")])),
            ({"order_id": "4", "brusher": ""}, ("g4", [Path("d.jpg")])),
        ]

        buckets = bucket_assignments_by_brusher(assignments)

        self.assertEqual([brusher for brusher, _ in buckets], ["唐杨", "小蝴蝶", ""])
        self.assertEqual([item[0]["order_id"] for item in buckets[0][1]], ["1", "3"])
        self.assertEqual([item[0]["order_id"] for item in buckets[1][1]], ["2"])
        self.assertEqual([item[0]["order_id"] for item in buckets[2][1]], ["4"])

    def test_date_suffix_for_filename_requires_single_date(self):
        self.assertEqual(date_suffix_for_filename([{"order_date_key": "20260514"}]), "20260514")

    def test_plan_group_batches_splits_all_dates_and_advances_rotation(self):
        records = [
            {"order_id": "o1", "order_date_key": "20260514"},
            {"order_id": "o2", "order_date_key": "20260514"},
            {"order_id": "o3", "order_date_key": "20260515"},
        ]
        groups = [
            ("1", [Path("1a.jpg"), Path("1b.jpg"), Path("1c.jpg"), Path("1d.jpg"), Path("1e.jpg")]),
            ("2", [Path("2a.jpg"), Path("2b.jpg"), Path("2c.jpg"), Path("2d.jpg"), Path("2e.jpg")]),
            ("3", [Path("3a.jpg"), Path("3b.jpg"), Path("3c.jpg"), Path("3d.jpg"), Path("3e.jpg")]),
            ("4", [Path("4a.jpg"), Path("4b.jpg"), Path("4c.jpg"), Path("4d.jpg"), Path("4e.jpg")]),
        ]

        batches, cursor_after = plan_group_batches(
            records=records,
            groups=groups,
            start_cursor=1,
            images_per_group=5,
            allow_total_shortage=0,
        )

        self.assertEqual([batch["date_key"] for batch in batches], ["20260514", "20260515"])
        self.assertEqual([name for name, _ in batches[0]["groups"]], ["2", "3"])
        self.assertEqual([name for name, _ in batches[1]["groups"]], ["4"])
        self.assertEqual(cursor_after, 0)

    def test_verify_group_image_counts_accepts_more_than_three_images_without_shortage(self):
        groups = [
            ("1", [Path("1.jpg"), Path("2.jpg"), Path("3.jpg"), Path("4.jpg")]),
        ]

        verify_group_image_counts(groups, images_per_group=5, allow_total_shortage=0)

    def test_verify_group_image_counts_rejects_three_images(self):
        groups = [
            ("1", [Path("1.jpg"), Path("2.jpg"), Path("3.jpg")]),
        ]

        with self.assertRaisesRegex(SystemExit, "图片不足"):
            verify_group_image_counts(groups, images_per_group=5, allow_total_shortage=0)

    def test_grouped_sources_prefers_group_dirs_without_flat_fallback(self):
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as td:
            tmp_path = Path(td)
            (tmp_path / "1").mkdir()
            for name in ["a.jpg", "b.jpg", "c.jpg"]:
                (tmp_path / "1" / name).write_text("x", encoding="utf-8")
            for name in ["loose1.jpg", "loose2.jpg", "loose3.jpg", "loose4.jpg", "loose5.jpg"]:
                (tmp_path / name).write_text("x", encoding="utf-8")

            groups = grouped_sources(tmp_path)

            self.assertEqual([name for name, _ in groups], ["1"])

    def test_plan_group_batches_fails_when_group_count_is_insufficient(self):
        records = [
            {"order_id": "o1", "order_date_key": "20260514"},
            {"order_id": "o2", "order_date_key": "20260514"},
        ]
        groups = [
            ("1", [Path("1a.jpg"), Path("1b.jpg"), Path("1c.jpg"), Path("1d.jpg"), Path("1e.jpg")]),
        ]

        with self.assertRaisesRegex(SystemExit, "分组不足"):
            plan_group_batches(
                records=records,
                groups=groups,
                start_cursor=0,
                images_per_group=5,
                allow_total_shortage=0,
            )

    def test_plan_group_batches_validates_only_selected_groups(self):
        records = [
            {"order_id": "o1", "order_date_key": "20260514"},
        ]
        groups = [
            ("1", [Path("1a.jpg"), Path("1b.jpg"), Path("1c.jpg")]),
            ("2", [Path("2a.jpg"), Path("2b.jpg"), Path("2c.jpg"), Path("2d.jpg")]),
        ]

        batches, cursor_after = plan_group_batches(
            records=records,
            groups=groups,
            start_cursor=1,
            images_per_group=5,
            allow_total_shortage=0,
        )

        self.assertEqual([name for name, _ in batches[0]["groups"]], ["2"])
        self.assertEqual(cursor_after, 0)

    def test_main_routes_to_workflow_without_packaging_or_patching(self):
        # 薄 wrapper 只透传参数给 workflow：结构上不 import 打包/回写能力，故无从直连。
        calls: list[list[str]] = []
        with (
            patch.object(sys, "argv", ["buyer_show", "--buyer-show-path", "/tmp/show", "--model", "AQA-12D-838", "--dry-run"]),
            patch.object(task_entry, "_run_workflow", side_effect=lambda args: calls.append(list(args)) or 0, create=True),
        ):
            result = task_entry.main()

        self.assertEqual(result, 0)
        self.assertEqual(calls, [["buyer_show", "--buyer-show-path", "/tmp/show", "--model", "AQA-12D-838", "--dry-run"]])
        self.assertFalse(hasattr(task_entry, "package_zip"))
        self.assertFalse(hasattr(task_entry, "patch_workbook"))

    def test_patch_workbook_updates_register_without_creating_backup(self):
        from tempfile import TemporaryDirectory

        with TemporaryDirectory() as td:
            tmp_path = Path(td)
            workbook_path = tmp_path / "register.xlsx"
            backup_dir = tmp_path / "backups"
            wb = Workbook()
            ws = wb.active
            ws["A3"] = "O1"
            ws["B3"] = "商品A"
            wb.save(workbook_path)

            with patch.object(buyer_show, "get_path", side_effect=lambda key: backup_dir if key == "backup_dir" else tmp_path):
                backup, verify = buyer_show.patch_workbook(
                    workbook_path,
                    [{"row": 3, "order_id": "O1", "name": "商品A"}],
                    {"订单编号": 0, "名称": 1},
                )

            self.assertIsNone(backup)
            self.assertEqual(list(backup_dir.glob("*")), [])
            self.assertEqual(verify[0]["status"], "是")


if __name__ == "__main__":
    unittest.main()
