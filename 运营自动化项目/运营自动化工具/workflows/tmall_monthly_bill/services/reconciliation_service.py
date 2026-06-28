from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def _normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def _coerce_cell_value(value: object) -> object:
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return ""
    return value


def _replace_sheet(workbook: Workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        workbook.remove(workbook[sheet_name])
    return workbook.create_sheet(sheet_name)


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = "" if value is None else str(value).strip()
    if len(text) < 10:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _filter_rows_by_period(
    header: list[str],
    rows: list[list[object]],
    *,
    period_start: str | date | None,
    period_end: str | date | None,
) -> list[list[object]]:
    if period_start is None and period_end is None:
        return rows
    start = _to_date(period_start)
    end = _to_date(period_end)
    if start is None or end is None:
        raise ValueError("对账单账期无效")
    try:
        period_index = header.index("账单周期")
    except ValueError as exc:
        raise ValueError("对账单列表 未找到账单周期列") from exc

    filtered: list[list[object]] = []
    for row in rows:
        if period_index >= len(row):
            continue
        period_text = "" if row[period_index] is None else str(row[period_index]).strip()
        if "~" not in period_text:
            continue
        row_start_text, row_end_text = (part.strip() for part in period_text.split("~", 1))
        row_start = _to_date(row_start_text)
        row_end = _to_date(row_end_text)
        if row_start is None or row_end is None:
            continue
        if row_end < start or row_start > end:
            continue
        filtered.append(row)
    return filtered


def _set_column_widths(worksheet, header: list[str], rows: list[list[object]]) -> None:
    sample_rows = rows[:300]
    for column_index, title in enumerate(header, start=1):
        values = [title]
        for row in sample_rows:
            value = row[column_index - 1] if column_index - 1 < len(row) else None
            values.append("" if value is None else str(value))
        max_length = max((len(value) for value in values), default=10)
        worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = min(max(max_length + 2, 10), 28)


def _read_first_sheet(path: Path) -> tuple[list[str], list[list[object]]]:
    source = load_workbook(path, data_only=True)
    worksheet = source[source.sheetnames[0]]
    header = [_normalize_header(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]
    rows: list[list[object]] = []
    for row_index in range(2, worksheet.max_row + 1):
        row = [_coerce_cell_value(worksheet.cell(row_index, column).value) for column in range(1, worksheet.max_column + 1)]
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return header, rows


def write_reconciliation_sheet(
    workbook: Workbook,
    source_path: Path,
    sheet_name: str = "对账单列表",
    *,
    period_start: str | date | None = None,
    period_end: str | date | None = None,
) -> None:
    header, rows = _read_first_sheet(source_path)
    rows = _filter_rows_by_period(rows=rows, header=header, period_start=period_start, period_end=period_end)
    worksheet = _replace_sheet(workbook, sheet_name)
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    _set_column_widths(worksheet, header, rows)
