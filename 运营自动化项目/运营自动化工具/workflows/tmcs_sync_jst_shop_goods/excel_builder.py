from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook


IMPORT_HEADERS = [
    "线上款式编码",
    "线上商品编码",
    "线上国标码",
    "平台店铺款式编码",
    "平台店铺商品编码",
    "原始商品编码",
    "线上商品名称",
    "线上颜色规格",
    "商品标识",
]

FAILED_HEADERS = ["platform_item_id", "platform_sku_id", "supplier_goods_id", "merchant_goods_code", "reason"]

# 猫超商品对应关系导入表（落桌面）的列，顺序对齐用户提供的模板
CORRESPONDENCE_HEADERS = [
    "平台店铺款式编码",
    "平台店铺商品编码",
    "线上商品名称",
    "线上颜色规格",
    "线上款式编码",
    "线上商品编码",
    "对应商品编码",
]

CORRESPONDENCE_FILE_NAME = "猫超商品对应关系导入表.xlsx"


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _norm_code(value: Any) -> str:
    """与「更新猫超商品列表」条码规则一致的归一化：去空格、去尾部 .0、转大写。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text.endswith(".0"):
        text = text[:-2]
    return text.upper()


def load_jst_code_pool(master_path: str | Path) -> tuple[dict[str, str], list[str]]:
    """从《聚水潭商品资料（最新）》的「商品编码」列构建匹配池（归一化精确表 + 去重列表）。"""
    from openpyxl import load_workbook

    path = Path(master_path)
    if not path.exists():
        raise FileNotFoundError(f"未找到聚水潭商品资料主数据：{path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [_text(cell) for cell in next(rows, [])]
    if "商品编码" not in header:
        workbook.close()
        raise ValueError(f"聚水潭商品资料缺少「商品编码」列：{path}")
    code_idx = header.index("商品编码")
    exact_map: dict[str, str] = {}
    normalized_codes: list[str] = []
    seen: set[str] = set()
    for row in rows:
        code = _norm_code(row[code_idx] if code_idx < len(row) else None)
        if not code:
            continue
        exact_map.setdefault(code, code)
        if code not in seen:
            normalized_codes.append(code)
            seen.add(code)
    workbook.close()
    return exact_map, normalized_codes


def match_corresponding_code(value: Any, exact_map: dict[str, str], normalized_codes: list[str]) -> str:
    """复刻条码规则：归一化后精确匹配 > 唯一模糊匹配（互为子串）；多义/未命中保留归一化原值。"""
    normalized = _norm_code(value)
    if not normalized:
        return ""
    exact = exact_map.get(normalized)
    if exact is not None:
        return exact
    fuzzy = [code for code in normalized_codes if normalized in code or code in normalized]
    if len(fuzzy) == 1:
        return fuzzy[0]
    return normalized


def build_correspondence_rows(
    import_rows: list[dict[str, str]],
    exact_map: dict[str, str],
    normalized_codes: list[str],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in import_rows:
        online_code = _text(row.get("线上商品编码"))
        rows.append(
            {
                "平台店铺款式编码": _text(row.get("平台店铺款式编码")),
                "平台店铺商品编码": _text(row.get("平台店铺商品编码")),
                "线上商品名称": _text(row.get("线上商品名称")),
                "线上颜色规格": _text(row.get("线上颜色规格")),
                "线上款式编码": _text(row.get("线上款式编码")),
                "线上商品编码": online_code,
                "对应商品编码": match_corresponding_code(online_code, exact_map, normalized_codes),
            }
        )
    return rows


def build_correspondence_workbook(
    *,
    import_rows: list[dict[str, str]],
    master_path: str | Path,
    output_dir: str | Path,
    file_name: str = CORRESPONDENCE_FILE_NAME,
) -> dict[str, str | int]:
    exact_map, normalized_codes = load_jst_code_pool(master_path)
    rows = build_correspondence_rows(import_rows, exact_map, normalized_codes)
    output_path = Path(output_dir) / file_name
    _write_workbook(output_path, CORRESPONDENCE_HEADERS, rows)
    matched = sum(1 for r in rows if r["对应商品编码"] and r["对应商品编码"] != _norm_code(r["线上商品编码"]))
    return {
        "correspondence_path": str(output_path),
        "correspondence_rows": len(rows),
        "matched_rows": matched,
    }


def build_rows(*, requested_item_ids: list[str], stock_rows: list[dict[str, Any]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    import_rows: list[dict[str, str]] = []
    failures: list[dict[str, str]] = []
    returned_items: set[str] = set()

    for row in stock_rows:
        platform_item_id = _text(row.get("platform_item_id"))
        platform_sku_id = _text(row.get("platform_sku_id"))
        supplier_goods_id = _text(row.get("supplier_goods_id"))
        merchant_goods_code = _text(row.get("merchant_goods_code"))
        if platform_item_id:
            returned_items.add(platform_item_id)

        reason = ""
        if not platform_item_id:
            reason = "平台商品ID为空"
        elif not supplier_goods_id:
            reason = "供应商货品ID为空"
        elif not merchant_goods_code:
            reason = "商家货品编码为空"

        if reason:
            failures.append(
                {
                    "platform_item_id": platform_item_id,
                    "platform_sku_id": platform_sku_id,
                    "supplier_goods_id": supplier_goods_id,
                    "merchant_goods_code": merchant_goods_code,
                    "reason": reason,
                }
            )
            continue

        import_rows.append(
            {
                "线上款式编码": platform_item_id,
                "线上商品编码": merchant_goods_code,
                "线上国标码": "",
                "平台店铺款式编码": platform_item_id,
                "平台店铺商品编码": supplier_goods_id,
                "原始商品编码": merchant_goods_code,
                "线上商品名称": "",
                "线上颜色规格": "",
                "商品标识": "Retail",
            }
        )

    for item_id in requested_item_ids:
        if item_id not in returned_items:
            failures.append(
                {
                    "platform_item_id": item_id,
                    "platform_sku_id": "",
                    "supplier_goods_id": "",
                    "merchant_goods_code": "",
                    "reason": "猫超未返回数据",
                }
            )

    return import_rows, failures


def _write_workbook(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "导入数据"
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_text(row.get(header)) for header in headers])
    for row in worksheet.iter_rows():
        for cell in row:
            cell.number_format = "@"
            if cell.value is None:
                cell.value = ""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def build_import_workbooks(
    *,
    import_rows: list[dict[str, str]],
    failures: list[dict[str, str]],
    output_dir: str | Path,
    timestamp: str,
) -> dict[str, str | int | None]:
    output_path = Path(output_dir)
    import_path = output_path / f"jst_shop_goods_import_{timestamp}.xlsx"
    failed_path = output_path / f"failed_items_{timestamp}.xlsx" if failures else None
    _write_workbook(import_path, IMPORT_HEADERS, import_rows)
    if failed_path:
        _write_workbook(failed_path, FAILED_HEADERS, failures)
    return {
        "import_path": str(import_path),
        "failed_path": str(failed_path) if failed_path else None,
        "import_rows": len(import_rows),
        "failed_rows": len(failures),
    }
