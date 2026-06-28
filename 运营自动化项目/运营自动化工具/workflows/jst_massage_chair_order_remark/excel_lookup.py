from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PRODUCT_CODE_HEADERS = ("商品编码", "产品编码", "编码", "product_code", "sku_code")
PRODUCT_NAME_HEADERS = ("商品名称", "产品名称", "品名", "名称", "product_name")


def _normalize(value: Any) -> str:
    return str(value or "").strip()


def _find_header(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    normalized = {_normalize(header).lower(): header for header in headers}
    for candidate in candidates:
        found = normalized.get(candidate.lower())
        if found:
            return found
    for header in headers:
        lowered = _normalize(header).lower()
        if any(candidate.lower() in lowered for candidate in candidates):
            return header
    return None


def load_massage_chair_mapping(path: str | Path) -> dict[str, str]:
    source = Path(path).expanduser().resolve()
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError(f"按摩椅资料表为空：{source}") from exc

    headers = [_normalize(value) for value in header_row]
    code_header = _find_header(headers, PRODUCT_CODE_HEADERS)
    name_header = _find_header(headers, PRODUCT_NAME_HEADERS)
    if not code_header or not name_header:
        raise ValueError(f"按摩椅资料表缺少商品编码或商品名称字段：{source}")

    code_idx = headers.index(code_header)
    name_idx = headers.index(name_header)
    mapping: dict[str, str] = {}
    for row in rows:
        code = _normalize(row[code_idx] if code_idx < len(row) else "")
        name = _normalize(row[name_idx] if name_idx < len(row) else "")
        if code and name:
            mapping[code] = name
    return mapping
