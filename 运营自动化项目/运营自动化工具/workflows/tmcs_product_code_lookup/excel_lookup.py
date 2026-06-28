"""猫超商品列表的本地读取与模糊匹配（纯业务逻辑，不接触任何平台）。

只读 Excel、按「商品上下架状态=上架」过滤、对用户输入的商品型号做多字段模糊匹配，
输出猫超商品编码等字段。绝不修改原始 Excel，绝不请求猫超/聚水潭后台。
"""

from __future__ import annotations

import json
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook, load_workbook


# ── 字段候选（按优先级，命中第一个存在的表头即采用，不硬猜）─────────────────────
# 必需字段：缺任意一个都无法完成查询，会停止并报出「缺失字段」。
STATUS_HEADER_CANDIDATES = ("商品上下架状态", "上下架状态", "状态")
PRODUCT_CODE_HEADER_CANDIDATES = ("商品编码", "猫超商品编码", "商品ID", "平台商品ID", "item_id")
SKU_HEADER_CANDIDATES = ("SKU编码", "平台SKUID", "SKU ID", "平台SKU编码")
BARCODE_HEADER_CANDIDATES = ("条码", "商品条码", "barcode")
NAME_HEADER_CANDIDATES = ("商品名称", "标题", "商品标题")
# 可选字段：缺失不影响主流程（匹配会回退到 条码/商品名称/SKU编码）。
# 品牌候选额外纳入真实导出表里的「淘系品牌名称 / 自营品牌名称」（值形如「AUX/奥克斯」）。
BRAND_HEADER_CANDIDATES = ("品牌", "品牌名称", "淘系品牌名称", "自营品牌名称")
MODEL_HEADER_CANDIDATES = ("产品型号", "型号", "规格型号", "货品型号")

# logical_field -> (候选表头, 是否必需)
_FIELD_SPEC: dict[str, tuple[tuple[str, ...], bool]] = {
    "status": (STATUS_HEADER_CANDIDATES, True),
    "product_code": (PRODUCT_CODE_HEADER_CANDIDATES, True),
    "sku_code": (SKU_HEADER_CANDIDATES, True),
    "barcode": (BARCODE_HEADER_CANDIDATES, True),
    "product_name": (NAME_HEADER_CANDIDATES, True),
    "brand": (BRAND_HEADER_CANDIDATES, False),
    "model": (MODEL_HEADER_CANDIDATES, False),
}

# 用于「中文缺失提示」的字段名。
_FIELD_LABELS = {
    "status": "商品上下架状态",
    "product_code": "猫超商品编码",
    "sku_code": "SKU编码",
    "barcode": "条码",
    "product_name": "商品名称",
    "brand": "品牌",
    "model": "产品型号",
}

ONLINE_STATUS_VALUES = ("上架",)
# 参与型号模糊匹配的字段（按业务语义：型号 > 条码 > SKU编码 > 商品名称）。
_MATCH_FIELDS = ("model", "barcode", "sku_code", "product_name")


class FieldResolutionError(ValueError):
    """必需字段在表头里找不到时抛出，附带清晰的缺失字段说明。"""


def _norm(value) -> str:
    return str(value if value is not None else "").strip()


def _header_map(sheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for column, raw_value in enumerate(header_row):
        header = _norm(raw_value)
        if header and header not in headers:
            headers[header] = column
    return headers


def resolve_fields(headers: dict[str, int]) -> tuple[dict[str, int], list[str]]:
    """把逻辑字段映射到表头列号。返回 (resolved, missing_required_labels)。

    resolved 只含命中的字段；缺失的必需字段会进 missing 列表（供调用方停止并提示）。
    """
    resolved: dict[str, int] = {}
    missing: list[str] = []
    for field, (candidates, required) in _FIELD_SPEC.items():
        column = next((headers[name] for name in candidates if name in headers), None)
        if column is None:
            if required:
                missing.append(_FIELD_LABELS[field])
            continue
        resolved[field] = column
    return resolved, missing


def load_online_products(path: Path) -> tuple[list[dict[str, str]], dict[str, int]]:
    """读取猫超商品列表，自动识别表头，只保留「商品上下架状态=上架」的行。

    返回 (products, resolved_fields)。products 中每行是逻辑字段 -> 字符串值。
    必需字段缺失时抛 FieldResolutionError，不硬猜。
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        headers = _header_map(sheet)
        resolved, missing = resolve_fields(headers)
        if missing:
            raise FieldResolutionError(
                "猫超商品列表缺少必需字段：" + "、".join(missing)
                + f"。实际表头：{', '.join(headers.keys())}"
            )
        status_col = resolved["status"]
        products: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            status = _norm(row[status_col] if status_col < len(row) else None)
            if status not in ONLINE_STATUS_VALUES:
                continue
            record = {
                field: _norm(row[col] if col < len(row) else None)
                for field, col in resolved.items()
            }
            products.append(record)
    finally:
        workbook.close()
    return products, resolved


def _ratio(query: str, text: str) -> float:
    """单字段相似度：相等=1.0，子串包含给高分，否则用序列相似度。"""
    q = query.strip().lower()
    t = text.strip().lower()
    if not q or not t:
        return 0.0
    if q == t:
        return 1.0
    if q in t:
        return round(0.9 + 0.1 * (len(q) / len(t)), 4)
    if t in q:
        return 0.8
    return round(SequenceMatcher(None, q, t).ratio(), 4)


def _model_score(query: str, product: dict[str, str]) -> float:
    return max((_ratio(query, product.get(field, "")) for field in _MATCH_FIELDS), default=0.0)


def _brand_matches(query: str, product: dict[str, str]) -> bool:
    """品牌过滤：优先用品牌字段；品牌字段缺失时回退到商品名称包含判断。"""
    bq = query.strip().lower()
    if not bq:
        return True
    brand_text = product.get("brand", "")
    if brand_text:
        return bq in brand_text.lower()
    return bq in product.get("product_name", "").lower()


def _dedupe_key(product: dict[str, str], dedupe_by: str) -> tuple:
    """去重键。默认按猫超商品编码（同一商品编码只出现一次）；

    dedupe_by="sku" 时退回到 (商品编码, SKU编码, 条码) 粒度，便于看每个 SKU 明细。
    商品编码为空时一律退回完整粒度，避免把不相关的空编码行误并成一条。
    """
    product_code = product.get("product_code", "")
    if dedupe_by == "product_code" and product_code:
        return ("product_code", product_code)
    return ("full", product_code, product.get("sku_code", ""), product.get("barcode", ""))


def fuzzy_match_products(
    products: list[dict[str, str]],
    *,
    model: str,
    brand: str | None = None,
    min_score: float = 0.5,
    limit: int = 10,
    dedupe_by: str = "product_code",
) -> list[dict]:
    """对上架商品做模糊匹配，去重并按匹配分数从高到低输出。

    dedupe_by：默认 "product_code"，同一猫超商品编码只保留匹配分最高的一条；
    传 "sku" 则按 SKU 粒度保留全部明细。
    """
    matched: dict[tuple, dict] = {}
    for product in products:
        if brand and not _brand_matches(brand, product):
            continue
        score = _model_score(model, product)
        if score < min_score:
            continue
        key = _dedupe_key(product, dedupe_by)
        result = {
            "product_code": product.get("product_code", ""),
            "sku_code": product.get("sku_code", ""),
            "barcode": product.get("barcode", ""),
            "brand": product.get("brand", ""),
            "model": product.get("model", ""),
            "product_name": product.get("product_name", ""),
            "match_score": round(score, 4),
        }
        prior = matched.get(key)
        if prior is None or result["match_score"] > prior["match_score"]:
            matched[key] = result
    ordered = sorted(
        matched.values(),
        key=lambda item: (-item["match_score"], item["product_code"], item["sku_code"]),
    )
    return ordered[:limit]


def write_result_json(path: Path, payload: dict) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def write_result_excel(path: Path, payload: dict) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "匹配结果"
    columns = ["猫超商品编码", "SKU编码", "条码", "品牌", "产品型号", "商品名称", "匹配分数"]
    sheet.append(columns)
    for item in payload.get("results", []):
        sheet.append(
            [
                item.get("product_code", ""),
                item.get("sku_code", ""),
                item.get("barcode", ""),
                item.get("brand", ""),
                item.get("model", ""),
                item.get("product_name", ""),
                item.get("match_score", ""),
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path
