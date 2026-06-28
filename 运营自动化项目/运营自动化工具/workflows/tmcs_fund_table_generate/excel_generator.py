from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font


SHEET_NAME = "店铺资金"
FUND_TABLE_HEADERS = [
    "运营",
    "不要修改店铺名，可以删减掉其他店铺名",
    "店铺编码",
    "平台",
    "年费",
    "保证金",
    "活动保证金余额",
    "账户余额",
    "待收货款",
    "推广账户余额",
    "营销账户余额",
    "支付宝余额",
    "备用金//微信余额",
    "银行卡余额",
    "提现对公户（店铺专用的一般户+基本户）",
    "预付款",
    "合计（扣除保证金+推广账户）",
    "x月货款预付",
    "项目合计可周转金额",
]

DEFAULT_ROW = [
    "国清",
    "（猫超）福安市启明工贸有限公司（国清）",
    "12633507",
    "天猫超市",
    None,
    50000,
    0,
    None,
    None,
    None,
    None,
    None,
    0,
    0,
    None,
    0,
    None,
    None,
    None,
]


@dataclass(frozen=True)
class FundTableResult:
    output_file: Path
    formula_check_result: dict[str, bool]


def _embed_in_cell(sheet, image_path: Path, cell: str, *, target_width: int = 320) -> tuple[int, int]:
    """把截图嵌入指定单元格（左上角锚定该格），仅缩小不放大，返回最终像素宽高。"""
    image = XLImage(str(image_path))
    if image.width and image.width > target_width:
        ratio = target_width / float(image.width)
        image.width = target_width
        image.height = int(image.height * ratio)
    sheet.add_image(image, cell)
    return int(image.width or target_width), int(image.height or 0)


def verify_fund_table(output_file: str | Path) -> dict[str, bool]:
    workbook = load_workbook(output_file)
    try:
        sheet = workbook[SHEET_NAME]
        return {
            "Q2": sheet["Q2"].value == "=H2+I2+L2+M2+N2+O2+P2",
            "S2": sheet["S2"].value == "=Q2-R2",
        }
    finally:
        workbook.close()


def generate_fund_table(
    *,
    output_file: str | Path,
    month: str,
    receivable_amount: float,
    promotion_balance: float,
    receivable_screenshot: str | Path,
    promotion_screenshot: str | Path,
    reserve_balance: float = 0,
    bank_card_balance: float = 0,
) -> FundTableResult:
    output_path = Path(output_file).expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    receivable_path = Path(receivable_screenshot).expanduser()
    promotion_path = Path(promotion_screenshot).expanduser()
    if not receivable_path.is_file():
        raise FileNotFoundError(f"待收货款截图不存在：{receivable_path}")
    if not promotion_path.is_file():
        raise FileNotFoundError(f"推广账户余额截图不存在：{promotion_path}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(FUND_TABLE_HEADERS)
    row = list(DEFAULT_ROW)
    row[8] = float(receivable_amount)
    row[9] = float(promotion_balance)
    row[12] = float(reserve_balance)
    row[13] = float(bank_card_balance)
    row[16] = "=H2+I2+L2+M2+N2+O2+P2"
    row[18] = "=Q2-R2"
    sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for index, header in enumerate(FUND_TABLE_HEADERS, start=1):
        width = min(max(len(header) + 2, 10), 26)
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row_index in range(2, 3):
        for col in range(5, 20):
            sheet.cell(row_index, col).number_format = "#,##0.00"

    # 截图凭证直接嵌入对应数据下一行同列单元格：待收货款 → I3、推广账户余额 → J3
    receivable_w, receivable_h = _embed_in_cell(sheet, receivable_path, "I3")
    promotion_w, promotion_h = _embed_in_cell(sheet, promotion_path, "J3")
    # 撑开 I/J 列宽与第 3 行行高，让图片视觉上落在单元格内（Excel 列宽≈像素/7，行高用磅，1px≈0.75pt）
    sheet.column_dimensions["I"].width = max(sheet.column_dimensions["I"].width or 0, receivable_w / 7.0)
    sheet.column_dimensions["J"].width = max(sheet.column_dimensions["J"].width or 0, promotion_w / 7.0)
    sheet.row_dimensions[3].height = max(receivable_h, promotion_h) * 0.75 or 15

    workbook.save(output_path)
    workbook.close()
    formula_check_result = verify_fund_table(output_path)
    return FundTableResult(output_file=output_path, formula_check_result=formula_check_result)
