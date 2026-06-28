#!/usr/bin/env python3
"""公司网盘下架产品上架数据的纯业务逻辑：自然语言解析、源目录定位、素材选取/复制、
聚水潭匹配、上架 Excel 生成与产出校验。

NAS 挂载/品牌·类目常量/编码归一化来自 workflows.company_nas_common.nas；本模块平台无关，
不做 workflow 编排。company_nas_listing workflow 的 steps.py 编排本模块。
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import shutil
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from core.config_loader import get_path  # noqa: E402
from workflows.company_nas_common.nas import (  # noqa: E402
    BRAND_FOLDERS,
    DEFAULT_NAS_MOUNT,
    NAS_CATEGORIES,
    NAS_MOUNT_NAME,
    NAS_URL,
    SKIP_NAMES,
    active_nas_mount,
    is_mounted,
    mount_nas,
    nas_product_root,
    normalize_code_text,
    run_command,
    unmount_nas,
)


PRODUCT_LIBRARY = get_path("nas_product_library_dir")
JST_WORKBOOK = get_path("jst_product_master_file")
NAS_INDEX_PATH = get_path("nas_index_json")

# WebDAV 列目录是延迟瓶颈，并发抓取目录/文件；过高会被群晖限流。
NAS_SCAN_WORKERS = 12
# 索引超过该天数视为可能过期：仅告警提示，不阻断（命中失败仍会回退实时遍历）。
NAS_INDEX_STALE_DAYS = 7

NATURAL_TEXT_PREFIXES = (
    "从公司网盘下载",
    "公司网盘下载产品",
    "公司网盘下载",
    "NAS产品资料下载",
    "下载产品资料并生成上架数据",
)

TARGET_BRAND_DIRS = {
    "奥克斯": PRODUCT_LIBRARY / "奥克斯产品",
    "苏泊尔": PRODUCT_LIBRARY / "苏泊尔产品",
}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
SKIP_EXTS = {".psd", ".mp4", ".mov", ".m4v", ".avi", ".db"}
NON_PRODUCT_WORDS = ("包材", "配件", "说明书", "网购箱", "彩盒", "泡沫", "适配器", "按摩头")
TITLE_NOISE_WORDS = ("优质", "爆款", "正品", "新款", "厂家", "批发", "专用", "适用")
WHITE_TRANSPARENT_ALIASES = {
    "白底透明",
    "白底图透明图",
    "透明图白底图",
    "白底透明图",
    "jpgpng",
}

# `company_nas_listing` is the single source of truth for NAS material selection.
# Docs and skills should summarize these rules instead of restating their own copies.
MAIN_IMAGE_PARENT_ALIASES = {"主图", "主附图"}
# 主图父目录下细分：纯「主图」单独成夹（→主图）；副主图/功能性主图/功能主图归「副图」。
MAIN_IMAGE_PRIMARY_CHILD_ALIASES = {"主图"}
MAIN_IMAGE_SECONDARY_CHILD_ALIASES = {"副主图", "功能性主图", "功能主图"}
MAIN_IMAGE_CHILD_ALIASES = MAIN_IMAGE_PRIMARY_CHILD_ALIASES | MAIN_IMAGE_SECONDARY_CHILD_ALIASES
SKU_DIR_ALIASES = {"sku"}
DETAIL_DIR_ALIASES = {"详情切片", "详情图"}
SCENE_DIR_ALIASES = {"场景图", "场景"}
SIZE_TOKENS_800 = ("800", "800x800", "800-800", "800x1200", "800x1000")
BLOCKED_PATH_KEYWORDS = ("视频", "旧版", "旧（厂家详情）")
DETAIL_DIR_NAME = "详情切片"
SCENE_DIR_NAME = "场景图"
MAIN_IMAGE_TARGET_DIR = "主图"
SECONDARY_IMAGE_TARGET_DIR = "副图"
SKU_TARGET_DIR = "sku"

CATEGORY_TITLE_PROFILES = {
    "按摩椅": {
        "功能": ["全身", "全自动", "智能", "太空舱", "零重力", "电动", "多功能", "家用"],
        "属性": ["3D", "4D", "AI", "语音", "热敷", "气囊", "SL导轨", "免安装"],
        "人群": ["老人", "父母", "办公室", "家用"],
    },
    "按摩靠垫": {
        "功能": ["腰背", "颈部", "揉捏", "热敷", "多功能", "智能", "家用"],
        "属性": ["全身", "靠背", "坐垫", "电动", "便携"],
        "人群": ["老人", "父母", "办公室", "车载"],
    },
    "按摩床垫": {
        "功能": ["全身", "揉捏", "热敷", "智能", "多功能", "家用"],
        "属性": ["床垫", "电动", "便携", "舒适", "折叠"],
        "人群": ["老人", "父母", "办公室", "家庭"],
    },
    "足疗机": {
        "功能": ["足底", "揉捏", "热敷", "气囊", "全自动", "电动", "恒温"],
        "属性": ["小腿", "脚底", "滚轮", "家用", "智能", "多功能"],
        "人群": ["老人", "父母", "办公室", "男女"],
    },
    "足浴盆": {
        "功能": ["泡脚", "恒温", "加热", "按摩", "全自动", "电动", "冲浪"],
        "属性": ["深桶", "折叠", "家用", "智能", "排水", "多功能"],
        "人群": ["老人", "父母", "家庭", "送礼"],
    },
    "护眼仪": {
        "功能": ["眼部", "热敷", "按摩", "气囊", "震动", "智能", "护眼"],
        "属性": ["蓝牙", "折叠", "便携", "恒温", "多模式", "充电"],
        "人群": ["学生", "办公", "睡眠", "男女"],
    },
    "筋膜枪": {
        "功能": ["深层", "肌肉", "放松", "按摩", "静音", "多档", "震动"],
        "属性": ["便携", "迷你", "长续航", "专业", "充电"],
        "人群": ["运动", "健身", "男女", "办公室"],
    },
    "护腰带": {
        "功能": ["腰部", "热敷", "按摩", "支撑", "护腰", "保暖"],
        "属性": ["充电", "可调节", "家用", "智能"],
        "人群": ["老人", "久坐", "男女", "办公室"],
    },
}

OUTPUT_COLUMNS = [
    "产品型号",
    "颜色",
    "天猫搜索标题（30字限制）",
    "SKU命名",
    "卖点1",
    "卖点2",
    "卖点3",
    "天猫搜索标题（15字）",
    "商品编码",
    "商品重量kg",
    "长cm",
    "宽cm",
    "高cm",
    "体积立方米",
    "总库存",
    "可用数",
    "匹配商品名称",
    "匹配备注",
]


@dataclass(frozen=True)
class ModelSpec:
    raw: str
    path_text: str
    manual_code: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="从公司 NAS 下载产品资料并生成上架数据.xlsx")
    parser.add_argument("--text", help="自然语言触发文本，例如：从公司网盘下载奥克斯足疗机AQA-JT-RFY06")
    parser.add_argument("--brand", help="品牌，例如：奥克斯、苏泊尔")
    parser.add_argument("--category", help="产品类别，例如：按摩靠垫、按摩椅、足浴盆")
    parser.add_argument("--models", nargs="*", default=[], help="型号列表；子目录用 / 或 \\ 表示，例如 AQA-12D-K10\\雾霾蓝")
    parser.add_argument("--models-file", help="型号文本文件，每行一个型号")
    parser.add_argument("--target-root", default=None, help="覆盖默认目标根目录")
    parser.add_argument("--jst-workbook", default=str(JST_WORKBOOK), help="聚水潭商品资料路径")
    parser.add_argument("--include-buyer-show", action="store_true", help="下载买家秀；默认不下载")
    parser.add_argument("--keep-mounted", action="store_true", help="执行结束后不卸载 NAS")
    parser.add_argument("--no-replace", action="store_true", help="不清理目标型号目录，直接增量复制")
    parser.add_argument("--skip-excel", action="store_true", help="只下载资料，不生成 Excel")
    parser.add_argument("--dry-run", action="store_true", help="只预览源目录、目标目录和预计文件数，不复制")
    return resolve_args(parser.parse_args())


def parse_natural_text(text: str) -> dict[str, Any]:
    normalized = re.sub(r"\s+", " ", text).strip()
    if not normalized:
        return {}
    brand = next((item for item in BRAND_FOLDERS if item in normalized), "")
    if not brand:
        return {}

    start = normalized.find(brand)
    remainder = normalized[start + len(brand):].strip(" ，,、")
    for prefix in NATURAL_TEXT_PREFIXES:
        if remainder.startswith(prefix):
            remainder = remainder[len(prefix):].strip(" ，,、")

    compact = re.sub(r"\s+", "", remainder)
    category = next((item for item in sorted(NAS_CATEGORIES, key=len, reverse=True) if compact.startswith(item)), "")
    if not category:
        return {"brand": brand}

    models_text = remainder
    if models_text.startswith(category):
        models_text = models_text[len(category):]
    else:
        models_text = re.sub(rf"^\s*{re.escape(category)}\s*", "", models_text, count=1)
    models_text = models_text.strip(" ，,、")
    models = [part.strip() for part in re.split(r"[，,、；;]+", models_text) if part.strip()]
    if len(models) == 1 and " " in models_text:
        models = [part.strip() for part in models_text.split() if part.strip()]
    return {"brand": brand, "category": category, "models": models}


def resolve_args(args: argparse.Namespace) -> argparse.Namespace:
    parsed = parse_natural_text(args.text or "")
    if not args.brand and parsed.get("brand"):
        args.brand = str(parsed["brand"])
    if not args.category and parsed.get("category"):
        args.category = str(parsed["category"])
    if not args.models and not args.models_file and parsed.get("models"):
        args.models = list(parsed["models"])

    missing = []
    if not args.brand:
        missing.append("--brand 或 --text 中的品牌")
    if not args.category:
        missing.append("--category 或 --text 中的类目")
    if not args.models and not args.models_file:
        missing.append("--models/--models-file 或 --text 中的型号")
    if missing:
        raise SystemExit(
            "公司 NAS 参数不足，缺少："
            + "、".join(missing)
            + "\n示例：python3 run.py \"从公司网盘下载奥克斯足疗机AQA-JT-RFY06\" --dry-run"
            + "\n或：python3 run.py company_nas_listing --brand 奥克斯 --category 足疗机 --models AQA-JT-RFY06 --dry-run"
        )
    return args


def normalize_model(raw: str) -> tuple[str, list[str]]:
    text = raw.strip().strip("\"'")
    parts = [p for p in re.split(r"[\\/]+", text) if p]
    if not parts:
        raise SystemExit("型号不能为空")
    return parts[0], parts[1:]


def parse_model_spec(raw: str) -> ModelSpec:
    text = raw.strip().strip("\"'")
    match = re.search(r"^(.*?)[【\[]\s*([^\]】]+?)\s*[】\]]\s*$", text)
    if match:
        path_text = match.group(1).strip()
        manual_code = match.group(2).strip()
    else:
        path_text = text
        manual_code = ""
    if not path_text:
        raise SystemExit(f"型号不能为空：{raw}")
    return ModelSpec(raw=text, path_text=path_text, manual_code=manual_code)


def load_models(args: argparse.Namespace) -> list[ModelSpec]:
    models = list(args.models)
    if args.models_file:
        path = Path(args.models_file).expanduser()
        if not path.is_file():
            raise SystemExit(f"型号文件不存在：{path}")
        models.extend([line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()])
    if not models:
        raise SystemExit("请通过 --models 或 --models-file 提供型号")
    return [parse_model_spec(model) for model in models]


def brand_source_dir(brand: str, category: str) -> Path:
    folder = BRAND_FOLDERS.get(brand)
    if not folder:
        raise SystemExit(f"暂未配置品牌目录：{brand}")
    return nas_product_root() / folder / category


def target_base_dir(brand: str, category: str, override: str | None) -> Path:
    if override:
        return Path(override).expanduser()
    return TARGET_BRAND_DIRS.get(brand, PRODUCT_LIBRARY / f"{brand}产品") / category


def model_source(base: Path, path_text: str) -> tuple[str, Path]:
    model, subdirs = normalize_model(path_text)
    src = base / model
    for part in subdirs:
        src = src / part
    display = "/".join([model, *subdirs])
    return display, src


def load_nas_index() -> dict[str, Any] | None:
    if not NAS_INDEX_PATH.exists():
        return None
    try:
        return json.loads(NAS_INDEX_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def score_index_record(record: dict[str, Any], *, brand: str, category: str, model: str) -> int:
    if record.get("type") != "dir":
        return 0
    if str(record.get("brand") or "") != brand:
        return 0
    if str(record.get("category") or "") != category:
        return 0

    needle = normalize_code_text(model)
    folder = normalize_code_text(record.get("model_or_folder", ""))
    filename = normalize_code_text(Path(str(record.get("path") or "")).name)
    keywords = " ".join(normalize_code_text(item) for item in (record.get("keywords") or []))
    score = 0
    if needle and needle == folder:
        score += 100
    elif needle and needle == filename:
        score += 95
    elif needle and needle in folder:
        score += 70
    elif needle and needle in filename:
        score += 65
    elif needle and needle in keywords:
        score += 40
    if int(record.get("depth") or 0) == 3:
        score += 15
    return score


def indexed_model_source(brand: str, category: str, base: Path, path_text: str) -> tuple[str, Path, str]:
    display, fallback = model_source(base, path_text)
    _, subdirs = normalize_model(path_text)
    if subdirs and fallback.is_dir():
        return display, fallback, "direct_path"

    payload = load_nas_index()
    if not payload:
        return display, fallback, "fallback_no_index"

    model, subdirs = normalize_model(path_text)
    candidates = []
    for record in payload.get("records") or []:
        score = score_index_record(record, brand=brand, category=category, model=model)
        if score <= 0:
            continue
        src = Path(str(record.get("path") or ""))
        for part in subdirs:
            src = src / part
        if src.is_dir():
            candidates.append((score, src))

    if not candidates:
        return display, fallback, "fallback_index_miss"

    candidates.sort(key=lambda item: (-item[0], len(item[1].parts), str(item[1])))
    return display, candidates[0][1], "nas_index"


def model_target(base: Path, path_text: str) -> tuple[str, Path]:
    model, subdirs = normalize_model(path_text)
    dst = base / model
    for part in subdirs:
        dst = dst / part
    display = "/".join([model, *subdirs])
    return display, dst


def is_800(path: Path) -> bool:
    s = str(path).lower()
    return any(token in s for token in SIZE_TOKENS_800)


def normalize_dir_name(name: str) -> str:
    return re.sub(r"[\W_]+", "", name).lower()


def dir_matches(name: str, aliases: set[str]) -> bool:
    return normalize_dir_name(name) in aliases


ChildDirCache = dict[Path, list[Path]]

_CACHE_LOCK = threading.Lock()


class _LiveLister:
    """实时遍历真实挂载（WebDAV/SMB）。目录/文件读取走网络。"""

    def child_dirs(self, base: Path) -> list[Path]:
        try:
            children = list(base.iterdir())
        except OSError:
            return []
        dirs: list[Path] = []
        for child in children:
            try:
                if child.is_dir():
                    dirs.append(child)
            except OSError:
                continue
        return dirs

    def iter_files(self, start: Path, include_buyer_show: bool):
        for root, dirs, files in os.walk(start):
            dirs[:] = [d for d in dirs if not should_skip_path(Path(root) / d, include_buyer_show)]
            for name in files:
                yield Path(root) / name


class _IndexLister:
    """从全量索引（含文件层级）在内存中应答目录/文件遍历，零网络。

    选材规则与实时遍历完全共用 `selected_files`，本类只替换「列目录 / 列文件」两个
    原语，从而保证索引选材与实时选材结果一致，不产生第二套选材逻辑。
    """

    def __init__(self, dir_paths: set[str], file_paths: list[Path]) -> None:
        self._file_paths = file_paths
        self._children: dict[str, list[Path]] = {}
        for raw in dir_paths:
            parent = str(Path(raw).parent)
            self._children.setdefault(parent, []).append(Path(raw))

    def child_dirs(self, base: Path) -> list[Path]:
        return list(self._children.get(str(base), []))

    def iter_files(self, start: Path, include_buyer_show: bool):
        prefix = str(start)
        prefix_sep = prefix + os.sep
        for path in self._file_paths:
            text = str(path)
            if text == prefix or text.startswith(prefix_sep):
                yield path


_LIVE_LISTER = _LiveLister()


def _parallel_map(func, items):
    """并发执行 func（IO 延迟瓶颈友好），保持与输入一致的结果顺序。"""
    items = list(items)
    if not items:
        return []
    workers = min(NAS_SCAN_WORKERS, len(items))
    if workers <= 1:
        return [func(item) for item in items]
    with ThreadPoolExecutor(max_workers=workers) as executor:
        return list(executor.map(func, items))


def safe_child_dirs(
    base: Path,
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
    lister: Any = None,
) -> list[Path]:
    lister = lister or _LIVE_LISTER
    if cache is not None:
        with _CACHE_LOCK:
            if base in cache:
                return cache[base]
    dirs = [child for child in lister.child_dirs(base) if not should_skip_path(child, include_buyer_show)]
    if cache is not None:
        with _CACHE_LOCK:
            cache[base] = dirs
    return dirs


def iter_matching_child_dirs(
    base: Path,
    aliases: set[str],
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
    lister: Any = None,
) -> list[Path]:
    return [child for child in safe_child_dirs(base, include_buyer_show, cache, lister) if dir_matches(child.name, aliases)]


def material_roots(
    base: Path,
    include_buyer_show: bool,
    cache: ChildDirCache | None = None,
    lister: Any = None,
) -> list[Path]:
    roots = [base]
    roots.extend(safe_child_dirs(base, include_buyer_show, cache, lister))
    return roots


def iter_main_image_dirs(
    base: Path,
    include_buyer_show: bool,
    roots: list[Path] | None = None,
    cache: ChildDirCache | None = None,
    lister: Any = None,
) -> list[Path]:
    matches: list[Path] = []
    for root in roots or material_roots(base, include_buyer_show, cache, lister):
        for parent in iter_matching_child_dirs(root, MAIN_IMAGE_PARENT_ALIASES, include_buyer_show, cache, lister):
            matches.extend(iter_matching_child_dirs(parent, MAIN_IMAGE_CHILD_ALIASES, include_buyer_show, cache, lister))
    return matches


def iter_category_dirs(
    base: Path,
    aliases: set[str],
    include_buyer_show: bool,
    roots: list[Path] | None = None,
    cache: ChildDirCache | None = None,
    lister: Any = None,
) -> list[Path]:
    matches: list[Path] = []
    for root in roots or material_roots(base, include_buyer_show, cache, lister):
        matches.extend(iter_matching_child_dirs(root, aliases, include_buyer_show, cache, lister))
    return matches


def should_skip_path(path: Path, include_buyer_show: bool) -> bool:
    # This gate owns the global skip policy: system junk, PSD/video, old dirs,
    # and buyer-show exclusion unless the caller explicitly opts in.
    text = str(path)
    if path.name in SKIP_NAMES or path.suffix.lower() in SKIP_EXTS:
        return True
    if not include_buyer_show and "买家秀" in text:
        return True
    if any(keyword in text for keyword in BLOCKED_PATH_KEYWORDS):
        return True
    return False


def collect_under(
    base: Path,
    rel_dir: Path | str,
    include_buyer_show: bool,
    predicate=lambda p: True,
    lister: Any = None,
) -> list[Path]:
    lister = lister or _LIVE_LISTER
    start = base / rel_dir
    out: list[Path] = []
    for p in lister.iter_files(start, include_buyer_show):
        if should_skip_path(p, include_buyer_show):
            continue
        if p.suffix.lower() not in IMAGE_EXTS:
            continue
        if predicate(p):
            out.append(p)
    return out


def _is_jpg_800(p: Path) -> bool:
    return is_800(p) and p.suffix.lower() in {".jpg", ".jpeg"}


def selected_files(base: Path, include_buyer_show: bool, *, lister: Any = None) -> list[Path]:
    # Selection order is intentional and defines the canonical download contract:
    # main images/SKU/scene images use the 800-token rule; details only take
    # `790*`; white-transparent assets are flattened to the product root; buyer
    # show stays off unless `--include-buyer-show` is set.
    #
    # `lister` 抽象「列目录 / 列文件」两个原语：默认 `_LiveLister` 实时走网络；传入
    # `_IndexLister` 则从全量索引内存应答（零网络）。两者共用本函数的选材规则，结果一致。
    # 目录遍历用线程池并发（WebDAV 是延迟瓶颈，并发可显著缩短等待）。
    lister = lister or _LIVE_LISTER
    child_dir_cache: ChildDirCache = {}
    roots = material_roots(base, include_buyer_show, child_dir_cache, lister)
    # 并发预热各 root 的子目录列举，后续发现阶段从缓存读取。
    _parallel_map(lambda r: safe_child_dirs(r, include_buyer_show, child_dir_cache, lister), roots)

    # 按 canonical 顺序收集 (源目录, 文件筛选谓词)，谓词与历史实现逐一对应。
    tasks: list[tuple[Path, Any]] = []
    for start in iter_main_image_dirs(base, include_buyer_show, roots, child_dir_cache, lister):
        tasks.append((start, _is_jpg_800))
    for start in iter_category_dirs(base, SKU_DIR_ALIASES, include_buyer_show, roots, child_dir_cache, lister):
        tasks.append((start, _is_jpg_800))
    for detail in iter_category_dirs(base, DETAIL_DIR_ALIASES, include_buyer_show, roots, child_dir_cache, lister):
        # Keep both historical layouts:
        # 1. detail/790/*.jpg
        # 2. detail/*.jpg|*.gif with "790" in filename
        def is_detail_790(p: Path, detail_dir: Path = detail) -> bool:
            if "790" in p.name:
                return True
            try:
                rel_parts = p.relative_to(detail_dir).parts[:-1]
            except ValueError:
                return False
            return any(part.startswith("790") for part in rel_parts)

        tasks.append((detail, is_detail_790))
    for start in iter_category_dirs(base, WHITE_TRANSPARENT_ALIASES, include_buyer_show, roots, child_dir_cache, lister):
        tasks.append((start, is_800))
    for start in iter_category_dirs(base, SCENE_DIR_ALIASES, include_buyer_show, roots, child_dir_cache, lister):
        tasks.append((start, _is_jpg_800))

    # 并发收集各目录文件；ThreadPoolExecutor.map 保持与 tasks 一致的顺序。
    files: list[Path] = []
    for chunk in _parallel_map(lambda t: collect_under(base, t[0], include_buyer_show, t[1], lister), tasks):
        files += chunk

    if include_buyer_show:
        files += collect_under(base, "买家秀", include_buyer_show, lister=lister)

    seen: set[Path] = set()
    unique: list[Path] = []
    for p in files:
        if p not in seen:
            unique.append(p)
            seen.add(p)
    return unique


def build_index_lister(payload: dict[str, Any] | None) -> _IndexLister | None:
    """从索引 payload 构建内存遍历器；无文件层级记录时返回 None（调用方回退实时遍历）。"""
    if not payload:
        return None
    dir_paths: set[str] = set()
    file_paths: list[Path] = []
    for record in payload.get("records") or []:
        path = str(record.get("path") or "")
        if not path:
            continue
        if record.get("type") == "dir":
            dir_paths.add(path)
        elif record.get("type") == "file":
            file_paths.append(Path(path))
    if not file_paths:
        return None
    return _IndexLister(dir_paths, file_paths)


def index_freshness(payload: dict[str, Any] | None) -> dict[str, Any]:
    """返回索引时效信息：updated_at / age_days / stale。"""
    updated = (payload or {}).get("updated_at")
    info: dict[str, Any] = {"updated_at": updated, "age_days": None, "stale": True}
    if not updated:
        return info
    try:
        ts = datetime.fromisoformat(str(updated))
    except ValueError:
        return info
    age = (datetime.now() - ts).days
    info["age_days"] = age
    info["stale"] = age > NAS_INDEX_STALE_DAYS
    return info


def select_files_resolved(
    src: Path,
    include_buyer_show: bool,
    index_payload: dict[str, Any] | None = None,
) -> tuple[list[Path], str]:
    """优先用全量索引在内存中选材（零网络）；索引无命中（新增产品/未建全量索引/过期）
    则回退实时遍历。返回 (files, source)，source ∈ {"index", "live"}。"""
    index_lister = build_index_lister(index_payload)
    if index_lister is not None:
        files = selected_files(src, include_buyer_show, lister=index_lister)
        if files:
            return files, "index"
    return selected_files(src, include_buyer_show), "live"


def _main_image_target(child_name: str) -> Path | None:
    """主图父目录下的子目录名 → 目标文件夹：纯主图归「主图」，副/功能主图归「副图」。"""
    if dir_matches(child_name, MAIN_IMAGE_PRIMARY_CHILD_ALIASES):
        return Path(MAIN_IMAGE_TARGET_DIR)
    if dir_matches(child_name, MAIN_IMAGE_SECONDARY_CHILD_ALIASES):
        return Path(SECONDARY_IMAGE_TARGET_DIR)
    return None


def copy_relative_path(src: Path, item: Path) -> Path:
    rel = item.relative_to(src)
    alias_parts = [part for part in rel.parts[:-1] if dir_matches(part, WHITE_TRANSPARENT_ALIASES)]
    if alias_parts:
        return Path(item.name)
    if any(dir_matches(part, SCENE_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(SCENE_DIR_NAME) / item.name
    if any(dir_matches(part, DETAIL_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(DETAIL_DIR_NAME) / item.name
    if any(dir_matches(part, SKU_DIR_ALIASES) for part in rel.parts[:-1]):
        return Path(SKU_TARGET_DIR) / item.name
    for idx, part in enumerate(rel.parts[:-1]):
        if not dir_matches(part, MAIN_IMAGE_PARENT_ALIASES):
            continue
        next_idx = idx + 1
        if next_idx < len(rel.parts) - 1:
            target = _main_image_target(rel.parts[next_idx])
            if target is not None:
                return target / item.name
    if len(rel.parts) >= 2 and dir_matches(rel.parts[0], MAIN_IMAGE_PARENT_ALIASES):
        target = _main_image_target(rel.parts[1])
        if target is not None:
            return target / item.name
    return rel


def copy_product(src: Path, dst: Path, files: list[Path], replace: bool, dry_run: bool) -> tuple[int, list[str]]:
    if dry_run:
        return 0, []
    if replace and dst.exists():
        shutil.rmtree(dst)
    dst.mkdir(parents=True, exist_ok=True)
    copied = 0
    missing_files: list[str] = []
    for item in files:
        rel = copy_relative_path(src, item)
        out = dst / rel
        if not replace and out.exists():
            continue
        if out.exists():
            out = out.with_name(f"{out.stem}_{copied + 1}{out.suffix}")
        out.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(item, out)
        except FileNotFoundError as exc:
            missing_files.append(f"{item} :: {exc}")
            continue
        except OSError as exc:
            missing_files.append(f"{item} :: {exc}")
            continue
        copied += 1
    return copied, missing_files


def row_text(row: tuple[Any, ...], idx: dict[str, int], columns: list[str]) -> str:
    values = []
    for col in columns:
        i = idx.get(col)
        values.append(str(row[i] or "") if i is not None and i < len(row) else "")
    return " ".join(values)


def load_jst_rows(path: Path) -> tuple[list[str], list[tuple[Any, ...]]]:
    if not path.is_file():
        raise SystemExit(f"聚水潭商品资料不存在：{path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return headers, rows


def jst_cell(row: tuple[Any, ...], idx: dict[str, int], name: str) -> Any:
    i = idx.get(name)
    return row[i] if i is not None and i < len(row) else None


def model_tokens(display_model: str) -> list[str]:
    parts = [p for p in display_model.split("/") if p]
    tokens = [parts[0], parts[0].replace("-", "")]
    tokens.extend(parts[1:])
    return [t.lower() for t in tokens if t]


def match_jst(display_model: str, manual_code: str, headers: list[str], rows: list[tuple[Any, ...]]) -> tuple[tuple[Any, ...] | None, str]:
    idx = {h: i for i, h in enumerate(headers)}
    code_idx = idx.get("商品编码")
    if code_idx is None:
        return None, "聚水潭资料缺少【商品编码】字段，需人工确认"

    if manual_code:
        target = normalize_code_text(manual_code)
        matches = [row for row in rows if normalize_code_text(jst_cell(row, idx, "商品编码")) == target]
        if not matches:
            return None, f"指定商品编码未匹配到聚水潭资料：{manual_code}"
        if len(matches) > 1:
            return matches[0], f"指定商品编码匹配到多条，已取第一条：{manual_code}"
        return matches[0], f"按指定商品编码精确匹配：{manual_code}"

    tokens = [normalize_code_text(token) for token in model_tokens(display_model)]
    tokens = [token for token in tokens if token]
    candidates = []
    for row in rows:
        product_code = str(jst_cell(row, idx, "商品编码") or "").strip()
        if not product_code:
            continue
        norm_code = normalize_code_text(product_code)
        score = 0
        for token in tokens:
            if norm_code == token:
                score = max(score, 100)
            elif token and token in norm_code:
                score = max(score, 60 + len(token))
            elif norm_code and norm_code in token:
                score = max(score, 40 + len(norm_code))
        if score:
            candidates.append((score, row))
    if not candidates:
        return None, "未提供商品编码，且按商品编码字段未匹配到聚水潭资料，需人工确认"
    candidates.sort(key=lambda item: item[0], reverse=True)
    top_score = candidates[0][0]
    same = [row for score, row in candidates if score == top_score]
    if len(same) > 1:
        return same[0], f"按商品编码字段模糊匹配到多条，已取第一条；候选数 {len(same)}"
    return candidates[0][1], "按商品编码字段模糊匹配"


def unique_terms(terms: list[str], category: str) -> list[str]:
    seen = set()
    out = []
    for term in terms:
        word = str(term).strip()
        if not word or word in seen or word == category or any(noise in word for noise in TITLE_NOISE_WORDS):
            continue
        seen.add(word)
        out.append(word)
    return out


def extract_title_terms(display_model: str, match: tuple[Any, ...] | None, idx: dict[str, int], category: str) -> list[str]:
    fields = ["商品名称", "商品简称", "分类", "颜色及规格", "颜色", "规格", "备注", "虚拟分类", "款式编码", "商品编码"]
    text = row_text(match, idx, fields) if match else display_model
    profile_words = []
    for profile in CATEGORY_TITLE_PROFILES.values():
        profile_words.extend(profile["功能"])
        profile_words.extend(profile["属性"])
        profile_words.extend(profile["人群"])
    candidates = [word for word in profile_words if word and word in text]
    color = infer_color(display_model, match, idx)
    if color:
        candidates.append(color)
    return unique_terms(candidates, category)


def category_profile(category: str) -> dict[str, list[str]]:
    return CATEGORY_TITLE_PROFILES.get(
        category,
        {
            "功能": ["智能", "多功能", "家用", "电动", "便携", "舒适"],
            "属性": ["大功率", "轻便", "实用", "升级"],
            "人群": ["家庭", "老人", "办公室", "送礼"],
        },
    )


def append_title_terms(base: str, terms: list[str], max_len: int) -> str:
    title = base
    for term in unique_terms(terms, ""):
        if term in title:
            continue
        if len(title) + len(term) <= max_len:
            title += term
    return title


_TITLE_LIBRARY_CACHE: dict[str, list[tuple[str, str]]] | None = None


def _load_title_library() -> dict[str, list[tuple[str, str]]]:
    """读取「按摩器材爆款标题库」，按类目分组缓存 (品牌, 商品标题) 列表。

    标题库每个 sheet 对应一个类目，列含「类目 / 品牌 / 商品标题」。
    品牌大多为空，留作类目级回退。库缺失或读取失败时返回空表，
    调用方自动回退到规则生成标题。
    """
    global _TITLE_LIBRARY_CACHE
    if _TITLE_LIBRARY_CACHE is not None:
        return _TITLE_LIBRARY_CACHE

    library: dict[str, list[tuple[str, str]]] = {}
    try:
        path = get_path("massage_title_library_file")
    except KeyError:
        _TITLE_LIBRARY_CACHE = library
        return library
    if not path.exists():
        _TITLE_LIBRARY_CACHE = library
        return library

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        _TITLE_LIBRARY_CACHE = library
        return library

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            continue
        col = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        title_idx = col.get("商品标题")
        if title_idx is None:
            continue
        cat_idx = col.get("类目")
        brand_idx = col.get("品牌")
        for row in rows:
            if title_idx >= len(row):
                continue
            title = row[title_idx]
            if not title or not str(title).strip():
                continue
            category = str(row[cat_idx]).strip() if cat_idx is not None and cat_idx < len(row) and row[cat_idx] else ws.title.strip()
            brand = str(row[brand_idx]).strip() if brand_idx is not None and brand_idx < len(row) and row[brand_idx] else ""
            library.setdefault(category, []).append((brand, str(title).strip()))
    wb.close()

    _TITLE_LIBRARY_CACHE = library
    return library


def pick_library_title(category: str, brand: str) -> str | None:
    """从标题库按「类目 + 品牌」随机取一个商品标题。

    优先匹配类目且品牌一致的标题；无命中则回退到该类目下任意标题。
    类目缺失或无标题时返回 None，由调用方回退到规则生成。
    """
    library = _load_title_library()
    entries = library.get((category or "").strip())
    if not entries:
        return None
    brand = (brand or "").strip()
    candidates = [title for entry_brand, title in entries if brand and entry_brand == brand]
    if not candidates:
        candidates = [title for _, title in entries]
    if not candidates:
        return None
    return random.choice(candidates)


def build_search_title(display_model: str, brand: str, category: str, match: tuple[Any, ...] | None, idx: dict[str, int]) -> str:
    profile = category_profile(category)
    extracted = extract_title_terms(display_model, match, idx, category)
    priority_terms = [*profile["功能"], *extracted, *profile["属性"], *profile["人群"]]
    title = append_title_terms(f"{brand}{category}", priority_terms, 30)
    title = title.replace(f"{category}{category}", category)
    if title.count(category) > 1:
        first = title.find(category)
        title = title[: first + len(category)] + title[first + len(category) :].replace(category, "")
    if len(title) < 28:
        title = append_title_terms(title, [*profile["属性"], *profile["人群"], "旗舰款", "送礼"], 30)
    return title[:30]


def compact_title(title: str, length: int = 15) -> str:
    return title[:length] if len(title) >= length else fit_len(title, length)


def fit_len(text: str, length: int) -> str:
    if len(text) >= length:
        return text[:length]
    return text + ("款" * (length - len(text)))


def infer_color(display_model: str, match: tuple[Any, ...] | None, idx: dict[str, int]) -> str:
    parts = display_model.split("/")
    if len(parts) > 1:
        return parts[-1]
    if match:
        name = str(jst_cell(match, idx, "商品名称") or "")
        m = re.search(r"[（(【](.*?色).*?[）)】]", name)
        if m:
            return m.group(1)
    return ""


def infer_brand(match: tuple[Any, ...] | None, idx: dict[str, int], fallback: str) -> str:
    if match:
        brand = str(jst_cell(match, idx, "品牌") or "").strip()
        if brand:
            return brand
    return fallback


def selling_points(category: str) -> list[str]:
    if category == "按摩靠垫":
        return ["腰背按摩", "热敷舒压", "揉捏放松"]
    if category == "按摩床垫":
        return ["全身按摩", "热敷舒压", "家用便携"]
    if category == "按摩椅":
        return ["AI语控", "零重体验", "全身拉伸"]
    return ["智能操控", "热敷舒压", "家用便携"]


def listing_row(display_model: str, brand: str, category: str, match: tuple[Any, ...] | None, remark: str, headers: list[str]) -> list[Any]:
    idx = {h: i for i, h in enumerate(headers)}
    title_brand = infer_brand(match, idx, brand)
    title30 = pick_library_title(category, title_brand) or build_search_title(display_model, title_brand, category, match, idx)
    title15 = compact_title(title30)
    color = infer_color(display_model, match, idx)
    points = selling_points(category)
    sku = f"【旗舰款】{color or '标准款'}{points[0]}"
    if match is None:
        return [display_model, color, title30, sku, *points, title15, "", "", "", "", "", "", "", "", "", remark]

    volume = jst_cell(match, idx, "体积")
    try:
        volume_m3 = round(float(volume) / 1000000, 6) if volume not in (None, "") else ""
    except Exception:
        volume_m3 = ""
    return [
        display_model,
        color,
        title30,
        sku,
        *points,
        title15,
        jst_cell(match, idx, "商品编码") or "",
        jst_cell(match, idx, "重量") or "",
        jst_cell(match, idx, "长") or "",
        jst_cell(match, idx, "宽") or "",
        jst_cell(match, idx, "高") or "",
        volume_m3,
        jst_cell(match, idx, "总库存") or 0,
        jst_cell(match, idx, "可用数") or 0,
        jst_cell(match, idx, "商品名称") or "",
        remark,
    ]


def save_listing(path: Path, rows: list[list[Any]], title: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "上架数据"
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_COLUMNS))
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(OUTPUT_COLUMNS)):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [18, 10, 34, 28, 10, 10, 10, 22, 16, 12, 10, 10, 10, 12, 10, 10, 42, 26]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(OUTPUT_COLUMNS))}{ws.max_row}"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def validate_outputs(base: Path | list[Path], listing_files: list[Path], include_buyer_show: bool) -> dict[str, Any]:
    bases = base if isinstance(base, list) else [base]
    invalid_files = []
    for item_base in bases:
        for root, _, files in os.walk(item_base):
            for name in files:
                p = Path(root) / name
                if p.suffix.lower() in SKIP_EXTS or p.name in SKIP_NAMES:
                    invalid_files.append(str(p))
    buyer_show_dirs = []
    if not include_buyer_show:
        for item_base in bases:
            buyer_show_dirs.extend(str(p) for p in item_base.rglob("买家秀") if p.is_dir())

    rule_errors = []
    for listing_file in listing_files:
        if listing_file.exists():
            wb = load_workbook(listing_file, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[2]]
            ci = {h: i + 1 for i, h in enumerate(headers)}
            for row_num in range(3, ws.max_row + 1):
                title30 = ws.cell(row_num, ci["天猫搜索标题（30字限制）"]).value or ""
                title15 = ws.cell(row_num, ci["天猫搜索标题（15字）"]).value or ""
                # 标题改为从「按摩器材爆款标题库」按类目+品牌随机取，长度在库内自然分布（约 26-35 字），
                # 仅在标题缺失或异常超长时报错，不再强卡 28-30 字。
                if not title30 or len(title30) > 40:
                    rule_errors.append({"file": str(listing_file), "row": row_num, "field": "30字标题", "length": len(title30)})
                if len(title15) != 15:
                    rule_errors.append({"file": str(listing_file), "row": row_num, "field": "15字标题", "length": len(title15)})
                for field in ("卖点1", "卖点2", "卖点3"):
                    value = ws.cell(row_num, ci[field]).value or ""
                    if len(value) > 4:
                        rule_errors.append({"file": str(listing_file), "row": row_num, "field": field, "length": len(value)})

    return {
        "listing_files_exist": all(p.exists() for p in listing_files),
        "buyer_show_dirs": buyer_show_dirs,
        "invalid_files": invalid_files,
        "rule_errors": rule_errors,
    }
