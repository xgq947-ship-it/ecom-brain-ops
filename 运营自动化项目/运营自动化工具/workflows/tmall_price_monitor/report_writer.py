"""天猫商品价格监控——结果产出（Excel 给人看 / JSON 给 Hermes 调用）。"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from workflows.tmall_price_monitor.price_compare import STATUS_BELOW, summarize


EXCEL_HEADERS = (
    "商品ID",
    "商品标题名称",
    "条码",
    "聚水潭商品编码",
    "聚水潭商品名称",
    "淘系控价",
    "商品实时价格",
    "商品差价",
    "状态",
    "抓取时间",
    "截图路径",
)

_BASENAME = "天猫商品价格监控"
_BELOW_FILL = PatternFill(start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")


def timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def output_names(ts: str | None = None) -> tuple[str, str]:
    ts = ts or timestamp()
    return f"{_BASENAME}_{ts}.xlsx", f"{_BASENAME}_{ts}.json"


def _cell(value: Any) -> Any:
    return "" if value is None else value


def write_excel(path: str | Path, records: list[dict[str, Any]]) -> Path:
    target = Path(path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "价格监控"
    ws.append(list(EXCEL_HEADERS))
    for col, _ in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL

    for record in records:
        ws.append(
            [
                record.get("item_id", ""),
                record.get("title", ""),
                record.get("barcode", ""),
                record.get("jst_goods_code", ""),
                record.get("jst_goods_name", ""),
                _cell(record.get("taoxi_control_price")),
                _cell(record.get("realtime_price")),
                _cell(record.get("diff_price")),
                record.get("status", ""),
                record.get("captured_at", ""),
                record.get("screenshot_path", ""),
            ]
        )
        if record.get("status") == STATUS_BELOW:
            for col in range(1, len(EXCEL_HEADERS) + 1):
                ws.cell(row=ws.max_row, column=col).fill = _BELOW_FILL

    widths = (16, 36, 16, 16, 28, 12, 14, 12, 16, 20, 50)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    wb.save(target)
    return target


def build_json_payload(records: list[dict[str, Any]], *, dry_run: bool = False, source: str = "page") -> dict[str, Any]:
    below = [
        {
            "item_id": r["item_id"],
            "title": r["title"],
            "barcode": r.get("barcode", ""),
            "jst_goods_code": r.get("jst_goods_code", ""),
            "jst_goods_name": r.get("jst_goods_name", ""),
            "taoxi_control_price": r.get("taoxi_control_price"),
            "realtime_price": r["realtime_price"],
            "diff_price": r["diff_price"],
        }
        for r in records
        if r["status"] == STATUS_BELOW
    ]
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "dry_run": dry_run,
        "total": len(records),
        "below_control_count": len(below),
        "summary": summarize(records),
        "below_control": below,
        "items": records,
    }


def write_json(path: str | Path, payload: dict[str, Any]) -> Path:
    target = Path(path).expanduser()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return target
