"""淘系控价自动匹配（纯本地 Excel/CSV 读取，不接触任何平台）。

链路：
  天猫商品ID
    -> 【猫超商品列表】按商品ID匹配出【条码】
    -> 用【条码】匹配【聚水潭商品资料】的【商品编码】
    -> 读取该聚水潭行的【淘系控价】作为 control_price

一个天猫商品ID 可能对应多个条码（多 SKU），各自控价不同。为「低于控价」告警尽量
不误报，取所有有效控价中的**最小值**作为代表控价（页面价低于最低控价才判定违规）。

只读文件，绝不修改原始 Excel；缺文件/缺字段/匹配失败都给出明确中文原因，不抛裸 traceback。
"""

from __future__ import annotations

import csv
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from core.config_loader import get_path


# ── 字段候选（命中第一个存在的表头即采用）─────────────────────────────────────
# 猫超商品列表：匹配输入商品ID 的列（真实导出里平台商品ID落在「商品编码」）。
MAOCHAO_ITEM_ID_CANDIDATES = (
    "商品ID", "item_id", "id", "平台商品ID", "宝贝ID",
    "商品编码", "货品编码", "SKU编码",
)
MAOCHAO_BARCODE_CANDIDATES = ("条码", "商品条码", "barcode", "国标码")
MAOCHAO_NAME_CANDIDATES = ("商品名称", "标题", "商品标题")

# 聚水潭商品资料：商品编码 == 猫超条码；淘系控价作为 control_price。
JST_CODE_CANDIDATES = ("商品编码", "款式编码", "商品编号", "goods_code")
JST_CONTROL_CANDIDATES = ("淘系控价", "控价", "taoxi_control_price")
JST_NAME_CANDIDATES = ("商品名称", "商品简称", "名称")

# 文件名识别关键词。
MAOCHAO_FILE_KEYWORDS = ("猫超商品列表", "天猫超市商品列表", "商品列表导出")
JST_FILE_KEYWORDS = ("聚水潭商品资料", "JST商品资料", "商品资料")
SUPPORTED_EXTS = (".xlsx", ".xls", ".csv")

# 匹配状态。
STATUS_MATCHED = "matched"
STATUS_NO_BARCODE = "未找到猫超条码"
STATUS_NO_JST = "未找到聚水潭商品"
STATUS_NO_CONTROL = "控价为空"


class MappingError(Exception):
    """文件/字段级别的可读错误（缺文件、缺必需字段等）。"""


def _debug(msg: str) -> None:
    # 调试输出走 stderr，避免污染 workflow 的 stdout JSON。
    print(f"[控价匹配] {msg}", file=sys.stderr)


# ── 清洗 ──────────────────────────────────────────────────────────────────────

def _cell_to_str(value) -> str:
    """把单元格值安全转字符串，避免长数字被科学计数法破坏。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        # 整数值的 float（如 762065566026.0）转成纯整数串，避免 .0 与科学计数法。
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value).strip()


def clean_item_id(value) -> str:
    """商品ID清洗：转字符串、去空格、去结尾 .0、不被科学计数法破坏。"""
    text = _cell_to_str(value).strip().replace(" ", "")
    url_id = re.search(r"[?&]id=(\d+)", text)
    if url_id:
        return url_id.group(1)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def clean_barcode(value) -> str:
    """条码/商品编码清洗：转字符串、去空格、去结尾 .0、保留前导 0。"""
    text = _cell_to_str(value).strip().replace(" ", "")
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def clean_price(value) -> float | None:
    """价格清洗：去 ¥￥元 逗号空格，提取第一个有效数字；空/无数字返回 None。"""
    text = _cell_to_str(value)
    if not text:
        return None
    cleaned = (
        text.replace("¥", "").replace("￥", "").replace("元", "")
        .replace(",", "").replace("，", "").replace(" ", "")
    )
    match = re.search(r"\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    try:
        return round(float(match.group(0)), 2)
    except ValueError:
        return None


# ── 文件查找 ──────────────────────────────────────────────────────────────────

def _candidate_dirs() -> list[Path]:
    dirs: list[Path] = []

    def add(path) -> None:
        try:
            p = Path(path).expanduser()
        except Exception:
            return
        if p not in dirs:
            dirs.append(p)

    # 优先项目锚点目录（主数据），再到常见落盘目录。
    try:
        add(get_path("maochao_goods_master_file").parent)  # 主数据
    except Exception:
        pass
    try:
        add(get_path("downloads_dir"))
    except Exception:
        pass
    try:
        add(get_path("desktop_dir"))
    except Exception:
        pass
    business_root = Path(__file__).resolve().parents[2]
    add(business_root / "data")
    add(business_root / "input")
    add(business_root / "downloads")
    add(Path.home() / "Downloads")
    add(Path.home() / "Desktop")
    return dirs


def _find_latest_by_keywords(keywords: tuple[str, ...]) -> Path | None:
    best: tuple[float, Path] | None = None
    for directory in _candidate_dirs():
        if not directory.is_dir():
            continue
        try:
            entries = list(directory.iterdir())
        except OSError:
            continue
        for entry in entries:
            if not entry.is_file():
                continue
            if entry.suffix.lower() not in SUPPORTED_EXTS:
                continue
            if entry.name.startswith("~$"):  # Excel 临时锁文件
                continue
            if not any(kw in entry.name for kw in keywords):
                continue
            mtime = entry.stat().st_mtime
            if best is None or mtime > best[0]:
                best = (mtime, entry)
    return best[1] if best else None


def find_source_file(kind: str, *, override: str | Path | None = None) -> Path:
    """定位猫超商品列表 / 聚水潭商品资料文件。

    优先级：显式 override > 项目配置锚点（存在即用）> 按关键词在候选目录找最新。
    """
    if override:
        path = Path(override).expanduser()
        if not path.exists():
            raise MappingError(f"指定的文件不存在：{path}")
        return path

    if kind == "maochao":
        anchors = ("maochao_goods_master_file", "tmall_goods_master_file", "tmall_goods_import_file")
        keywords = MAOCHAO_FILE_KEYWORDS
        label = "猫超商品列表"
    elif kind == "jst":
        anchors = ("jst_product_master_file", "jst_product_file", "jst_product_import_file")
        keywords = JST_FILE_KEYWORDS
        label = "聚水潭商品资料"
    else:
        raise MappingError(f"未知文件类型：{kind}")

    for key in anchors:
        try:
            candidate = get_path(key)
        except Exception:
            continue
        if candidate and Path(candidate).exists():
            return Path(candidate)

    found = _find_latest_by_keywords(keywords)
    if found:
        return found
    raise MappingError(
        f"未找到{label}文件（关键词：{'/'.join(keywords)}）。"
        f"请把文件放到主数据/、~/Downloads/ 或 ~/Desktop/，支持 xlsx/xls/csv。"
    )


# ── 表读取 ────────────────────────────────────────────────────────────────────

def _read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    """读取 xlsx/xls/csv，统一返回 (headers, rows)，所有单元格清洗为字符串。"""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            raw = list(csv.reader(handle))
        if not raw:
            return [], []
        headers = [_cell_to_str(c) for c in raw[0]]
        rows = [[_cell_to_str(c) for c in line] for line in raw[1:]]
        return headers, rows

    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            raise MappingError("缺少 openpyxl，请先安装依赖。") from exc
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = wb[wb.sheetnames[0]]
            it = sheet.iter_rows(values_only=True)
            header_row = next(it, ())
            headers = [_cell_to_str(c) for c in header_row]
            rows = [[_cell_to_str(c) for c in row] for row in it]
        finally:
            wb.close()
        return headers, rows

    if suffix == ".xls":
        # 旧二进制 .xls 需要 xlrd（项目未装）；给出明确指引而非裸 traceback。
        try:
            import xlrd  # type: ignore
        except ModuleNotFoundError as exc:
            raise MappingError(
                f"暂不支持 .xls（缺 xlrd）：{path.name}。请用 Excel 另存为 .xlsx 后重试。"
            ) from exc
        book = xlrd.open_workbook(str(path))  # pragma: no cover - 环境无 xlrd
        sheet = book.sheet_by_index(0)
        headers = [_cell_to_str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
        rows = [
            [_cell_to_str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            for r in range(1, sheet.nrows)
        ]
        return headers, rows

    raise MappingError(f"不支持的文件类型：{path.suffix}（仅支持 xlsx/xls/csv）。")


def _resolve_col(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    index = {h: i for i, h in enumerate(headers) if h}
    for name in candidates:
        if name in index:
            return index[name]
    return None


def _resolve_cols(headers: list[str], candidates: tuple[str, ...]) -> list[int]:
    index = {h: i for i, h in enumerate(headers) if h}
    return [index[name] for name in candidates if name in index]


# ── 索引构建 ──────────────────────────────────────────────────────────────────

@dataclass
class _MaochaoIndex:
    by_item_id: dict[str, set[str]] = field(default_factory=dict)
    name_by_item_id: dict[str, str] = field(default_factory=dict)
    id_columns: list[str] = field(default_factory=list)
    barcode_column: str = ""


def load_maochao_index(path: Path) -> _MaochaoIndex:
    headers, rows = _read_rows(path)
    if not headers:
        raise MappingError(f"猫超商品列表为空或无表头：{path}")
    id_cols = _resolve_cols(headers, MAOCHAO_ITEM_ID_CANDIDATES)
    barcode_col = _resolve_col(headers, MAOCHAO_BARCODE_CANDIDATES)
    name_col = _resolve_col(headers, MAOCHAO_NAME_CANDIDATES)
    if not id_cols:
        raise MappingError(
            "猫超商品列表缺少商品ID列（候选："
            + "/".join(MAOCHAO_ITEM_ID_CANDIDATES)
            + f"）。实际表头：{', '.join(headers)}"
        )
    if barcode_col is None:
        raise MappingError(
            "猫超商品列表缺少条码列（候选：" + "/".join(MAOCHAO_BARCODE_CANDIDATES) + "）。"
        )

    idx = _MaochaoIndex(
        id_columns=[headers[c] for c in id_cols],
        barcode_column=headers[barcode_col],
    )
    for row in rows:
        barcode = clean_barcode(row[barcode_col]) if barcode_col < len(row) else ""
        name = _cell_to_str(row[name_col]) if (name_col is not None and name_col < len(row)) else ""
        for col in id_cols:
            if col >= len(row):
                continue
            item_id = clean_item_id(row[col])
            if not item_id:
                continue
            if barcode:
                idx.by_item_id.setdefault(item_id, set()).add(barcode)
            else:
                idx.by_item_id.setdefault(item_id, set())
            if name and item_id not in idx.name_by_item_id:
                idx.name_by_item_id[item_id] = name
    return idx


@dataclass
class _JstIndex:
    by_code: dict[str, tuple[float | None, str]] = field(default_factory=dict)
    code_column: str = ""
    control_column: str = ""


def load_jst_index(path: Path) -> _JstIndex:
    headers, rows = _read_rows(path)
    if not headers:
        raise MappingError(f"聚水潭商品资料为空或无表头：{path}")
    code_col = _resolve_col(headers, JST_CODE_CANDIDATES)
    control_col = _resolve_col(headers, JST_CONTROL_CANDIDATES)
    name_col = _resolve_col(headers, JST_NAME_CANDIDATES)
    if code_col is None:
        raise MappingError(
            "聚水潭商品资料缺少商品编码列（候选：" + "/".join(JST_CODE_CANDIDATES) + "）。"
        )
    if control_col is None:
        raise MappingError(
            "聚水潭商品资料缺少淘系控价列（候选：" + "/".join(JST_CONTROL_CANDIDATES) + "）。"
        )

    idx = _JstIndex(code_column=headers[code_col], control_column=headers[control_col])
    for row in rows:
        code = clean_barcode(row[code_col]) if code_col < len(row) else ""
        if not code:
            continue
        control = clean_price(row[control_col]) if control_col < len(row) else None
        name = _cell_to_str(row[name_col]) if (name_col is not None and name_col < len(row)) else ""
        existing = idx.by_code.get(code)
        # 同编码多行：优先保留有控价的一条。
        if existing is None or (existing[0] is None and control is not None):
            idx.by_code[code] = (control, name)
    return idx


# ── 解析器 ────────────────────────────────────────────────────────────────────

class ControlPriceResolver:
    """加载两张表各一次，对每个商品ID解析淘系控价。"""

    def __init__(
        self,
        *,
        maochao_path: str | Path | None = None,
        jst_path: str | Path | None = None,
        debug: bool = True,
    ) -> None:
        self.debug = debug
        self.maochao_path = find_source_file("maochao", override=maochao_path)
        self.jst_path = find_source_file("jst", override=jst_path)
        if self.debug:
            _debug(f"猫超商品列表文件：{self.maochao_path}")
            _debug(f"聚水潭商品资料文件：{self.jst_path}")
        self.maochao = load_maochao_index(self.maochao_path)
        self.jst = load_jst_index(self.jst_path)
        if self.debug:
            _debug(
                f"猫超索引：{len(self.maochao.by_item_id)} 个商品ID"
                f"（ID列={'/'.join(self.maochao.id_columns)}，条码列={self.maochao.barcode_column}）"
            )
            _debug(
                f"聚水潭索引：{len(self.jst.by_code)} 个商品编码"
                f"（控价列={self.jst.control_column}）"
            )

    def resolve(self, raw_item_id: str) -> dict:
        item_id = clean_item_id(raw_item_id)
        result = {
            "item_id": item_id,
            "barcode": "",
            "jst_goods_code": "",
            "jst_goods_name": "",
            "maochao_name": self.maochao.name_by_item_id.get(item_id, ""),
            "taoxi_control_price": None,
            "mapping_status": STATUS_NO_BARCODE,
            "matched_barcode_count": 0,
            "all_control_prices": [],
        }

        barcodes = sorted(self.maochao.by_item_id.get(item_id, set()))
        if not barcodes:
            if self.debug:
                _debug(f"{item_id} → 未在猫超商品列表匹配到商品ID/条码")
            return result

        result["matched_barcode_count"] = len(barcodes)
        if self.debug:
            _debug(f"{item_id} → 猫超条码：{', '.join(barcodes)}")

        candidates: list[tuple[str, float | None, str]] = []
        for bc in barcodes:
            j = self.jst.by_code.get(clean_barcode(bc))
            if j is not None:
                candidates.append((bc, j[0], j[1]))

        if not candidates:
            result["mapping_status"] = STATUS_NO_JST
            if self.debug:
                _debug(f"{item_id} → 条码均未匹配到聚水潭商品编码")
            return result

        valid = [(bc, ctrl, name) for bc, ctrl, name in candidates if ctrl is not None]
        result["all_control_prices"] = sorted({ctrl for _, ctrl, _ in valid})
        if not valid:
            bc, _, name = candidates[0]
            result.update({
                "barcode": bc,
                "jst_goods_code": bc,
                "jst_goods_name": name,
                "mapping_status": STATUS_NO_CONTROL,
            })
            if self.debug:
                _debug(f"{item_id} → 聚水潭匹配成功但淘系控价为空")
            return result

        # 多条码取最小控价作为代表控价（最保守，少误报「低于控价」）。
        bc, ctrl, name = min(valid, key=lambda t: t[1])
        result.update({
            "barcode": bc,
            "jst_goods_code": bc,
            "jst_goods_name": name,
            "taoxi_control_price": ctrl,
            "mapping_status": STATUS_MATCHED,
        })
        if self.debug:
            _debug(f"{item_id} → 淘系控价 {ctrl}（条码 {bc}，共 {len(valid)} 个有效控价）")
        return result


def get_control_price_by_item_id(
    item_id: str,
    *,
    resolver: ControlPriceResolver | None = None,
    maochao_path: str | Path | None = None,
    jst_path: str | Path | None = None,
) -> dict:
    """便捷入口：解析单个商品ID的淘系控价。

    批量场景请复用同一个 ControlPriceResolver，避免重复读表。
    """
    resolver = resolver or ControlPriceResolver(maochao_path=maochao_path, jst_path=jst_path)
    return resolver.resolve(item_id)
