"""猫超店铺商品销售分析 workflow 的 step handler。

分层边界：
- CSV 下载动作只在 Ops-Cli（`ops --json jst report product-profit export`）。业务层只通过
  clients.ops_cli_client.run_ops_json 调用，不写聚水潭 URL/Cookie/Token/Selector/Playwright/CDP。
- 本层只负责：获取 CSV 路径 → 调 csv_analyzer.analyze_sales_csv → 收集店铺款式编码 → 输出与 Artifact。

dry-run 安全点：
- fetch_sales_csv 向 Ops-Cli 透传 dry-run（不带 --execute），平台层返回 simulated=true，不真实导出。
- dry-run 且未提供 --use-local-file 时，无真实 CSV，analyze/write 步骤安全跳过（不报错）。
- write_outputs 在 dry-run 下不落盘结果文件。
"""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date, timedelta
from pathlib import Path

from clients.ops_cli_client import run_ops_json
from core.config_loader import get_path
from core.runtime import parse_workflow_args, Artifact, StepContext, failure_result, success_result

from workflows.jst_tmcs_shop_product_sales_analysis import csv_analyzer

DEFAULT_SHOP_NAME = "（猫超）福安市启明工贸有限公司"


def _last_month() -> str:
    first_of_this_month = date.today().replace(day=1)
    return (first_of_this_month - timedelta(days=1)).strftime("%Y-%m")


def _normalize_period_flags(namespace: argparse.Namespace) -> None:
    if namespace.days is not None:
        if namespace.start_date or namespace.end_date:
            raise ValueError("--days 不能和 --start-date/--end-date 同时使用")
        if namespace.days <= 0:
            raise ValueError("--days 必须大于 0")
        end = date.today()
        start = end - timedelta(days=namespace.days - 1)
        namespace.start_date = start.isoformat()
        namespace.end_date = end.isoformat()

    if namespace.start_date or namespace.end_date:
        if namespace.month:
            raise ValueError("--month 不能和 --start-date/--end-date 同时使用")
        if not namespace.start_date or not namespace.end_date:
            raise ValueError("--start-date 和 --end-date 必须同时提供")
        try:
            start = date.fromisoformat(namespace.start_date)
            end = date.fromisoformat(namespace.end_date)
        except ValueError as exc:
            raise ValueError("日期只支持 YYYY-MM-DD") from exc
        if start > end:
            raise ValueError("--start-date 不能晚于 --end-date")
        namespace.period_label = f"{start.isoformat()}_to_{end.isoformat()}"
        return

    if not namespace.month:
        namespace.month = _last_month()
    namespace.period_label = namespace.month


def _parse_flags(ctx: StepContext) -> argparse.Namespace:
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--month", default=None)
    parser.add_argument("--days", type=int, default=None)
    parser.add_argument("--start-date", default=None)
    parser.add_argument("--end-date", default=None)
    parser.add_argument("--shop-name", default=DEFAULT_SHOP_NAME)
    parser.add_argument("--use-local-file", default=None)
    parser.add_argument("--output", default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--execute", action="store_true")
    namespace = parse_workflow_args(parser, ctx.inputs.get("args") or [])
    namespace.dry_run = ctx.dry_run or namespace.dry_run
    _normalize_period_flags(namespace)
    return namespace


def check_inputs(ctx: StepContext):
    flags = _parse_flags(ctx)
    # 未指定 --output 时，默认生成分档位推广清单 Excel 到桌面。
    if not flags.output:
        flags.output = str(get_path("desktop_dir") / f"猫超店铺推广清单_{flags.period_label}.xlsx")
        ctx.state["output_defaulted"] = True
    ctx.state["flags"] = flags
    return success_result(
        outputs={
            "month": flags.month,
            "days": flags.days,
            "start_date": flags.start_date,
            "end_date": flags.end_date,
            "period_label": flags.period_label,
            "shop_name": flags.shop_name,
            "use_local_file": flags.use_local_file,
            "output": flags.output,
            "output_defaulted": ctx.state.get("output_defaulted", False),
            "dry_run": flags.dry_run,
            "execute": flags.execute,
        }
    )


def fetch_sales_csv(ctx: StepContext):
    flags = ctx.state["flags"]

    # 路径 1：本地 CSV，直接使用，不触平台。
    if flags.use_local_file:
        local = Path(str(flags.use_local_file)).expanduser()
        if not local.is_file():
            return failure_result(errors=[f"--use-local-file 指定的 CSV 不存在：{local}"])
        ctx.state["csv_path"] = str(local)
        ctx.state["csv_source"] = "local_file"
        return success_result(
            outputs={"csv_path": str(local), "source": "local_file", "downloaded": False}
        )

    # 路径 2：调用 Ops-Cli 导出（下载动作在平台层）。
    command = ["--json", "jst", "report", "product-profit", "export", "--shop-name", flags.shop_name]
    if flags.start_date and flags.end_date:
        command.extend(["--start-date", flags.start_date, "--end-date", flags.end_date])
    else:
        command.extend(["--month", flags.month])
    if flags.dry_run:
        command.append("--dry-run")
    else:
        command.append("--execute")

    try:
        payload = run_ops_json(command, interactive_recovery=not flags.dry_run)
    except RuntimeError as exc:
        if flags.dry_run:
            ctx.state["csv_path"] = None
            ctx.state["csv_source"] = "ops_unavailable"
            return success_result(outputs={"skipped": True, "reason": str(exc), "csv_path": None})
        return failure_result(errors=[f"Ops-Cli 导出失败：{exc}"])

    data = payload.get("data") if isinstance(payload, dict) else {}
    data = data if isinstance(data, dict) else {}
    csv_path = data.get("csv_path")
    ctx.state["ops_data"] = data
    ctx.state["csv_path"] = csv_path
    ctx.state["csv_source"] = "ops_export"

    if flags.dry_run or not csv_path:
        return success_result(
            outputs={
                "skipped": True,
                "reason": "dry-run：Ops-Cli 返回 simulated 预览，未真实导出 CSV" if flags.dry_run else "Ops-Cli 未返回 csv_path",
                "simulated": bool(data.get("simulated", flags.dry_run)),
                "shop_name": data.get("shop_name", flags.shop_name),
                "month": data.get("month", flags.month),
                "period": data.get("period"),
                "period_label": data.get("period_label", flags.period_label),
                "csv_path": None,
            }
        )
    return success_result(
        outputs={
            "csv_path": csv_path,
            "source": "ops_export",
            "downloaded": True,
            "month": data.get("month", flags.month),
            "period": data.get("period"),
            "period_label": data.get("period_label", flags.period_label),
        }
    )


def analyze_sales_csv(ctx: StepContext):
    csv_path = ctx.state.get("csv_path")
    if not csv_path:
        return success_result(
            outputs={"skipped": True, "reason": "无可分析的 CSV（dry-run 未导出且未提供 --use-local-file）"}
        )
    try:
        result = csv_analyzer.analyze_sales_csv(str(csv_path))
    except (FileNotFoundError, ValueError) as exc:
        return failure_result(errors=[str(exc)])

    ctx.state["analysis"] = result
    return success_result(
        outputs={
            "style_codes": result["style_codes"],
            "total_rows": result["total_rows"],
            "matched_rows": result["matched_rows"],
            "unique_style_code_count": result["unique_style_code_count"],
            "categories": result["categories"],
        }
    )


# 推广档位定义：(明细 key, 段标题, 推荐说明)
_PROMO_SECTIONS = [
    ("priority", "🏆 优先推广", "全站推广（优先）→ 万相台关键词推广（叠加），目标推广占比 4%~8%"),
    ("secondary", "✅ 次级推广", "全站推广小预算测试，日预算 100~200 元/品"),
    ("warning", "⚠️ 推广过高预警", "建议降低出价或缩减预算"),
    ("stop", "🛑 建议暂停推广", "推广中但经营亏损，建议暂停"),
]


def _write_promotion_xlsx(output_path: Path, analysis: dict, *, period_label: str, shop_name: str) -> None:
    """生成分档位推广清单 Excel（店铺款式编码 / 商品名称 / 利润率 / 销量 / 销售额 / 当前推广）。"""
    from openpyxl import Workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    details = analysis.get("details") or {}
    workbook = Workbook()
    ws = workbook.active
    ws.title = "推广清单"

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    pct = lambda v: f"{v * 100:.1f}%"

    columns = ["店铺款式编码", "商品名称", "利润率", "毛利率", "销量(件)", "销售额(元)", "均价(元)", "退款率", "当前推广(元)", "推广占比"]

    ws.append([f"猫超店铺推广SKU清单　{shop_name}　{period_label}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1).font = title_font
    ws.append([])

    for key, section_title, advice in _PROMO_SECTIONS:
        records = details.get(key) or []
        # 段标题行
        r = ws.max_row + 1
        ws.append([f"{section_title}（{len(records)} 个）　{advice}"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(columns))
        c = ws.cell(row=r, column=1)
        c.font = section_font
        c.fill = section_fill
        # 列头
        ws.append(columns)
        for col in range(1, len(columns) + 1):
            hc = ws.cell(row=ws.max_row, column=col)
            hc.font = header_font
            hc.fill = header_fill
        # 数据
        if records:
            for p in records:
                ws.append([
                    p["sku"], p["name"], pct(p["op_margin"]), pct(p["gross_margin"]),
                    round(p["qty"]), round(p["amt"], 2), round(p["avg_price"]),
                    pct(p["refund_rate"]), round(p["ad_consume"], 2), pct(p["ad_pct"]),
                ])
        else:
            ws.append(["（暂无符合条件商品）"])
        ws.append([])

    # 列宽
    widths = [18, 40, 9, 9, 9, 12, 9, 9, 12, 9]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 店铺款式编码列设为文本，避免长编码被当成数字/科学计数
    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if not isinstance(cell, MergedCell) and isinstance(cell.value, str):
                cell.number_format = "@"
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    workbook.save(output_path)


def _write_result_file(output_path: Path, analysis: dict, *, period_label: str, shop_name: str, month: str | None = None) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    if suffix == ".json":
        payload = {
            "shop_name": shop_name,
            "month": month,
            "period_label": period_label,
            "style_codes": analysis["style_codes"],
            "total_rows": analysis["total_rows"],
            "matched_rows": analysis["matched_rows"],
            "unique_style_code_count": analysis["unique_style_code_count"],
            "categories": analysis["categories"],
        }
        output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    elif suffix == ".xlsx":
        _write_promotion_xlsx(output_path, analysis, period_label=period_label, shop_name=shop_name)
    else:  # 默认 CSV
        with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["店铺款式编码"])
            for code in analysis["style_codes"]:
                writer.writerow([code])


def write_outputs(ctx: StepContext):
    flags = ctx.state["flags"]
    analysis = ctx.state.get("analysis")

    if not analysis:
        return success_result(outputs={"skipped": True, "reason": "无分析结果，无需写出"})

    if not flags.output:
        return success_result(
            outputs={
                "written": False,
                "reason": "未指定 --output，仅返回结果",
                "style_codes": analysis["style_codes"],
            }
        )

    output_path = Path(str(flags.output)).expanduser()
    if flags.dry_run:
        return success_result(
            outputs={"skipped": True, "reason": "dry-run：跳过写出结果文件", "planned_output": str(output_path)}
        )

    _write_result_file(output_path, analysis, period_label=flags.period_label, month=flags.month, shop_name=flags.shop_name)
    ctx.state["output_path"] = str(output_path)

    # 分析输出成功后，自动删除「我们下载的」原始 CSV（--use-local-file 你自己的文件不删）。
    deleted_source = None
    csv_path = ctx.state.get("csv_path")
    if (
        ctx.state.get("csv_source") == "ops_export"
        and output_path.exists()
        and csv_path
        and Path(str(csv_path)).exists()
    ):
        try:
            Path(str(csv_path)).unlink()
            deleted_source = str(csv_path)
            ctx.state["source_csv_deleted"] = True
        except OSError as exc:
            ctx.state["source_csv_delete_error"] = str(exc)

    return success_result(
        outputs={"written": True, "output_path": str(output_path), "deleted_source_csv": deleted_source}
    )


def collect_artifacts(ctx: StepContext):
    flags = ctx.state["flags"]
    analysis = ctx.state.get("analysis")
    csv_path = ctx.state.get("csv_path")
    output_path = ctx.state.get("output_path")
    artifacts: list[Artifact] = []

    source_deleted = ctx.state.get("source_csv_deleted", False)
    if csv_path and Path(str(csv_path)).exists():
        # 原始 CSV 仍在（如 --use-local-file，或删除失败）：正常登记
        artifacts.append(
            Artifact(
                type="csv",
                role="sales_source",
                name=Path(str(csv_path)).name,
                path=str(csv_path),
                platform="jst",
                month=flags.month,
                metadata={"period_label": flags.period_label},
            )
        )
    elif csv_path and source_deleted:
        # 已分析并删除的原始下载 CSV：保留一条溯源记录（文件已不在）
        artifacts.append(
            Artifact(
                type="csv",
                role="sales_source",
                name=Path(str(csv_path)).name,
                path=str(csv_path),
                platform="jst",
                month=flags.month,
                metadata={"deleted_after_analysis": True, "period_label": flags.period_label},
            )
        )
    if output_path and Path(str(output_path)).exists():
        artifacts.append(
            Artifact(
                type=Path(str(output_path)).suffix.lstrip(".") or "csv",
                role="style_code_output",
                name=Path(str(output_path)).name,
                path=str(output_path),
                platform="jst",
                month=flags.month,
                metadata={"period_label": flags.period_label},
            )
        )

    return success_result(
        outputs={
            "task": "jst_tmcs_shop_product_sales_analysis",
            "dry_run": flags.dry_run,
            "month": flags.month,
            "start_date": flags.start_date,
            "end_date": flags.end_date,
            "period_label": flags.period_label,
            "shop_name": flags.shop_name,
            "csv_path": csv_path,
            "csv_source": ctx.state.get("csv_source"),
            "source_csv_deleted": ctx.state.get("source_csv_deleted", False),
            "output_path": output_path,
            "style_codes": analysis["style_codes"] if analysis else [],
            "unique_style_code_count": analysis["unique_style_code_count"] if analysis else 0,
        },
        artifacts=artifacts,
    )
