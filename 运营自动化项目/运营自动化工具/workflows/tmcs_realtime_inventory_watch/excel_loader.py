"""猫超库存实时监测 — 本地 Excel 读取与字段识别。

只负责把三张本地 Excel（聚水潭商品资料、猫超库存明细、猫超商品列表）按候选字段名
读成统一的 dict 行，并完成数值清洗、去重。不碰平台，不写任何平台逻辑。

字段识别规则：每个逻辑字段给一组候选列名，按表头顺序取第一个命中的列；
若一个都识别不到，抛 FieldNotFoundError，附上实际表头，便于排错（不猜错列）。

数值清洗：去除千分位逗号、全角空格、单位等噪声后转 float；空值按 0 处理并记 warning。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


class FieldNotFoundError(ValueError):
    """需要的逻辑字段在表头里一个候选都识别不到。"""


# ── 候选字段名 ────────────────────────────────────────────────────────────────
# 表1 聚水潭商品资料
JST_PRODUCT_CODE = ("商品编码", "商品编号", "款式编码", "商家编码")
JST_ACTUAL_STOCK = ("实际库存", "实际库存数", "库存", "当前库存")
JST_ORDER_HOLD = ("订单占有", "订单占有数", "订单占用", "占有库存")
JST_BRAND = ("品牌", "品牌名称")

# 表2 猫超库存明细
TMCS_PLATFORM_SKU_ID = ("平台SKUID", "平台SKU ID", "平台skuId", "SKU编码", "skuId")
TMCS_DEDICATED_SELLABLE = ("专享现货库存可售量", "专享库存", "专享可售库存")
TMCS_SHARED_SELLABLE = ("共享现货库存可售量", "共享库存", "共享可售库存")
TMCS_WAREHOUSE_CODE = ("商家仓code", "商家仓CODE", "仓库编码", "warehouse_code", "storeCode")

# 表3 猫超商品列表
GOODS_SKU_CODE = ("SKU编码", "平台SKUID", "SKU ID")
GOODS_BARCODE = ("条码", "商品条码", "barcode")
GOODS_LISTING_STATUS = ("商品上下架状态", "上下架状态", "状态")
GOODS_NAME = ("商品名称", "商品名", "名称")


def _norm(text: Any) -> str:
    return "" if text is None else str(text).strip()


def norm_id(value: Any) -> str:
    """规范化 ID/编码类字段，用于跨表关联。

    Excel 把 13 位数字 ID 存成 float，str() 会得到 "5265333653202.0"，
    破坏与字符串 ID 的相等比较。这里把整数值的 float/int 还原为纯整数字符串，
    其余（含字母编码如 AUYZAMQST1202C01）按普通 strip 处理。
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).strip()
    # 形如 "5265333653202.0" 的纯数字 float 文本也还原为整数
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".")[0]
    return text


def read_rows(path: str | Path) -> tuple[list[str], list[dict[str, Any]]]:
    """读取 Excel 第一个 sheet，返回 (表头, 行 dict 列表)。"""
    file_path = Path(path).expanduser()
    if not file_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{file_path}")
    # 不用 read_only：部分平台导出的 <dimension> 元数据不准（如只标 A1），
    # read_only 模式会据此漏读列/行；非 read_only 会按实际单元格扫描，更稳。
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        workbook.close()
        return [], []
    headers = [_norm(cell) for cell in header_raw]
    records: list[dict[str, Any]] = []
    for raw in rows_iter:
        if raw is None or all(cell is None or _norm(cell) == "" for cell in raw):
            continue
        record = {headers[idx]: raw[idx] for idx in range(min(len(headers), len(raw)))}
        records.append(record)
    workbook.close()
    return headers, records


def resolve_field(headers: Iterable[str], candidates: Iterable[str], *, label: str) -> str:
    """在表头里按候选顺序找第一个命中的列名，找不到抛 FieldNotFoundError。"""
    header_set = {h for h in headers}
    for candidate in candidates:
        if candidate in header_set:
            return candidate
    raise FieldNotFoundError(
        f"无法识别字段「{label}」：候选 {list(candidates)} 均不在表头中。实际表头：{list(headers)}"
    )


_NUMERIC_NOISE = re.compile(r"[,，\s　]")


def clean_number(value: Any) -> tuple[float, bool]:
    """清洗为 float。返回 (数值, 是否为空/缺失按0处理)。

    空值/无法解析 -> (0.0, True)，调用方据此记 warning。
    """
    if value is None:
        return 0.0, True
    if isinstance(value, (int, float)):
        return float(value), False
    text = _NUMERIC_NOISE.sub("", str(value)).strip()
    if text == "":
        return 0.0, True
    # 去掉非数字尾巴（如单位），保留前导符号、数字、小数点
    match = re.match(r"^[-+]?\d*\.?\d+", text)
    if not match:
        return 0.0, True
    try:
        return float(match.group()), False
    except ValueError:
        return 0.0, True
