from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


PRIORITY_STYLE_HEADERS = ("优先推广店铺款式编码", "店铺款式编码", "平台店铺款式编码", "款式编码")
PRIORITY_ITEM_HEADERS = ("商品ID", "平台商品ID", "item_id", "Item ID")
PRIORITY_PRODUCT_HEADERS = ("商品编码", "产品编码", "货品编码", "SKU编码")
ACTIVE_STYLE_HEADERS = ("店铺款式编码", "平台店铺款式编码", "款式编码", "商品编码")
MASTER_STYLE_HEADERS = ("商品编码", "店铺款式编码", "平台店铺款式编码", "SKU编码", "货品编码")
MASTER_PRODUCT_HEADERS = ("商品编码", "货品编码", "平台商品编码")
MASTER_ITEM_HEADERS = ("商品编码", "商品ID", "平台商品ID", "item_id")


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_rows(path: Path) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _row_text(row: list[Any]) -> str:
    return " | ".join(normalize_cell(value) for value in row if normalize_cell(value))


def _find_header_indexes(header_row: list[Any], candidates: tuple[str, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = normalize_cell(value)
        if header:
            mapping[header] = index
    return mapping


def _pick_header_index(header_map: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        if candidate in header_map:
            return header_map[candidate]
    return None


def _looks_like_priority_header(row: list[Any]) -> bool:
    header_map = _find_header_indexes(row, PRIORITY_STYLE_HEADERS)
    return _pick_header_index(header_map, PRIORITY_STYLE_HEADERS) is not None


def load_priority_promotion_rows(path: Path) -> list[dict[str, str]]:
    rows = _load_rows(path)
    section_header_index: int | None = None
    in_priority_section = False

    for index, row in enumerate(rows):
        first_cell = normalize_cell(row[0] if row else "")
        if "优先推广" in first_cell:
            in_priority_section = True
            continue
        if in_priority_section and _looks_like_priority_header(row):
            section_header_index = index
            break

    if section_header_index is None:
        for index, row in enumerate(rows):
            if _looks_like_priority_header(row):
                section_header_index = index
                break

    if section_header_index is None:
        raise ValueError(
            "推广清单中无法识别优先推广字段，至少需要以下字段之一："
            + " / ".join(PRIORITY_STYLE_HEADERS)
        )

    header_row = rows[section_header_index]
    header_map = _find_header_indexes(header_row, PRIORITY_STYLE_HEADERS)
    style_index = _pick_header_index(header_map, PRIORITY_STYLE_HEADERS)
    if style_index is None:
        raise ValueError(
            "推广清单中无法识别优先推广字段，至少需要以下字段之一："
            + " / ".join(PRIORITY_STYLE_HEADERS)
        )

    parsed: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    for row in rows[section_header_index + 1 :]:
        first_cell = normalize_cell(row[0] if row else "")
        if not any(normalize_cell(value) for value in row):
            if parsed:
                break
            continue
        if first_cell.startswith(("✅", "⚠️", "🛑")) or "次级推广" in first_cell or "推广过高预警" in first_cell or "建议暂停" in first_cell:
            break
        style_code = normalize_cell(row[style_index] if style_index < len(row) else None)
        if not style_code:
            continue
        if style_code in seen_codes:
            continue
        seen_codes.add(style_code)
        item = {"store_style_code": style_code}
        for headers_group in (PRIORITY_ITEM_HEADERS, PRIORITY_PRODUCT_HEADERS):
            for header in headers_group:
                index = header_map.get(header)
                if index is None:
                    continue
                value = normalize_cell(row[index] if index < len(row) else None)
                if value:
                    item[header] = value
        parsed.append(item)
    return parsed


def load_active_style_codes(path: Path) -> list[str]:
    rows = _load_rows(path)
    if not rows:
        return []
    header_map = _find_header_indexes(rows[0], ACTIVE_STYLE_HEADERS)
    style_index = _pick_header_index(header_map, ACTIVE_STYLE_HEADERS)
    if style_index is None:
        raise ValueError(
            "正在推广商品列表中无法识别字段，至少需要以下字段之一："
            + " / ".join(ACTIVE_STYLE_HEADERS)
        )
    seen: set[str] = set()
    result: list[str] = []
    for row in rows[1:]:
        value = normalize_cell(row[style_index] if style_index < len(row) else None)
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def append_active_style_codes(path: Path, style_codes: list[str]) -> dict[str, Any]:
    normalized_codes: list[str] = []
    seen_input: set[str] = set()
    for raw_code in style_codes:
        code = normalize_cell(raw_code)
        if not code or code in seen_input:
            continue
        seen_input.add(code)
        normalized_codes.append(code)

    if not normalized_codes:
        return {"appended_codes": [], "appended_count": 0}

    workbook = load_workbook(path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            worksheet.append([ACTIVE_STYLE_HEADERS[0]])
            style_index = 0
        else:
            header_map = _find_header_indexes(list(rows[0]), ACTIVE_STYLE_HEADERS)
            style_index = _pick_header_index(header_map, ACTIVE_STYLE_HEADERS)
            if style_index is None:
                raise ValueError(
                    "正在推广商品列表中无法识别字段，至少需要以下字段之一："
                    + " / ".join(ACTIVE_STYLE_HEADERS)
                )

        existing = set(load_active_style_codes(path))
        appended_codes: list[str] = []
        for code in normalized_codes:
            if code in existing:
                continue
            row_values = [""] * (style_index + 1)
            row_values[style_index] = code
            worksheet.append(row_values)
            cell = worksheet.cell(row=worksheet.max_row, column=style_index + 1)
            cell.value = code
            cell.number_format = "@"
            existing.add(code)
            appended_codes.append(code)
        workbook.save(path)
        return {"appended_codes": appended_codes, "appended_count": len(appended_codes)}
    finally:
        workbook.close()


def load_tmcs_master_mapping(path: Path) -> dict[str, dict[str, str]]:
    rows = _load_rows(path)
    if not rows:
        return {}
    header_map = _find_header_indexes(rows[0], MASTER_STYLE_HEADERS)
    product_index = _pick_header_index(header_map, MASTER_PRODUCT_HEADERS)
    item_index = _pick_header_index(header_map, MASTER_ITEM_HEADERS)
    if product_index is None and item_index is None:
        raise ValueError(
            "猫超商品主表缺少商品编码字段，至少需要以下字段之一："
            + " / ".join(MASTER_ITEM_HEADERS)
        )

    style_indexes = [header_map[name] for name in MASTER_STYLE_HEADERS if name in header_map]
    mapping: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        product_code = normalize_cell(row[product_index] if product_index is not None and product_index < len(row) else None)
        item_id = normalize_cell(row[item_index] if item_index is not None and item_index < len(row) else None)
        if not product_code and item_id:
            product_code = item_id
        if not item_id and product_code:
            item_id = product_code
        if not item_id:
            continue
        record = {
            "product_code": product_code,
            "item_id": item_id,
        }
        for index in style_indexes:
            key = normalize_cell(row[index] if index < len(row) else None)
            if key and key not in mapping:
                mapping[key] = record
    return mapping


def write_result_json(path: Path, payload: dict[str, Any]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return path


def write_result_csv(path: Path, payload: dict[str, Any]) -> Path:
    import csv

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["group", "store_style_code", "product_code", "item_id", "skip_reason", "plan_name"])
        for group in ("priority_items", "skipped_items", "to_create_items", "created_items", "failed_items"):
            for item in payload.get(group, []):
                writer.writerow(
                    [
                        group,
                        item.get("store_style_code", ""),
                        item.get("product_code", ""),
                        item.get("item_id", ""),
                        item.get("skip_reason", item.get("reason", "")),
                        item.get("plan_name", ""),
                    ]
                )
    return path


def write_result_excel(path: Path, payload: dict[str, Any]) -> Path:
    workbook = Workbook()
    headers = ["store_style_code", "product_code", "item_id", "skip_reason", "plan_name", "daily_budget", "target_workflow", "status", "error"]
    sheet_specs = [
        ("priority_items", "priority_items"),
        ("skipped_items", "skipped_items"),
        ("to_create_items", "to_create_items"),
        ("created_items", "created_items"),
        ("failed_items", "failed_items"),
    ]
    first_sheet = True
    for key, title in sheet_specs:
        worksheet = workbook.active if first_sheet else workbook.create_sheet(title=title)
        worksheet.title = title
        worksheet.append(headers)
        for item in payload.get(key, []):
            worksheet.append([item.get(header, "") for header in headers])
        first_sheet = False
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path
